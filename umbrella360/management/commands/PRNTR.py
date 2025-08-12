import pandas as pd
from django.core.management.base import BaseCommand
from django.db.models import Avg, Sum, Count, Max, Min
from umbrella360.models import Empresa, Unidade, Viagem_Base, CheckPoint, Infrações
import os
from datetime import datetime

class Command(BaseCommand):
    help = 'Gera relatório Excel para a empresa CPBRASCELL'

    def handle(self, *args, **options):
        try:
            # Buscar a empresa CPBRACELL
            empresa = Empresa.objects.get(nome='CPBRACELL')
            self.stdout.write(f"Empresa encontrada: {empresa.nome}")
            
        except Empresa.DoesNotExist:
            self.stdout.write(self.style.ERROR("Empresa CPBRACELL não encontrada no banco de dados"))
            return

        # Criar nome do arquivo com timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"relatorio_CPBRACELL_{timestamp}.xlsx"
        
        # Preparar dados para as planilhas
        self.stdout.write("Coletando dados das unidades...")
        dados_unidades = self.coletar_dados_unidades(empresa)
        
        self.stdout.write("Coletando dados das viagens...")
        dados_viagens = self.coletar_dados_viagens(empresa)
        
        # Criar o arquivo Excel
        self.stdout.write("Gerando arquivo Excel...")
        with pd.ExcelWriter(nome_arquivo, engine='openpyxl') as writer:
            # Primeira planilha - Dados das Unidades
            df_unidades = pd.DataFrame(dados_unidades)
            df_unidades.to_excel(writer, sheet_name='Unidades', index=False)
            
            # Segunda planilha - Dados das Viagens por Período
            df_viagens = pd.DataFrame(dados_viagens)
            df_viagens.to_excel(writer, sheet_name='Viagens por Período', index=False)
            
            # Formatação básica das planilhas
            self.formatar_planilhas(writer)
        
        self.stdout.write(
            self.style.SUCCESS(
                f"Relatório gerado com sucesso: {nome_arquivo}\n"
                f"Total de unidades: {len(dados_unidades)}\n"
                f"Total de viagens: {len(dados_viagens)}"
            )
        )

    def coletar_dados_unidades(self, empresa):
        """Coleta dados das unidades da empresa com estatísticas"""
        unidades = Unidade.objects.filter(empresa=empresa).order_by('cls', 'id')
        dados = []
        
        for unidade in unidades:
            # Buscar viagens da unidade com filtros de eficiência
            viagens_unidade = Viagem_Base.objects.filter(
                unidade=unidade,
                Quilometragem_média__gte=1.0,
                Quilometragem_média__lte=4.0
            )
            
            # Calcular estatísticas
            stats = viagens_unidade.aggregate(
                total_viagens=Count('id'),
                total_quilometragem=Sum('quilometragem'),
                total_consumo=Sum('Consumido'),
                eficiencia_media=Avg('Quilometragem_média'),
                velocidade_media=Avg('Velocidade_média'),
                rpm_medio=Avg('RPM_médio'),
                temperatura_media=Avg('Temperatura_média'),
                total_emissoes=Sum('Emissões_CO2'),
                melhor_eficiencia=Max('Quilometragem_média'),
                pior_eficiencia=Min('Quilometragem_média'),
                ultimo_periodo=Max('período')
            )
            
            # Buscar dados de checkpoints
            checkpoints_stats = CheckPoint.objects.filter(unidade=unidade).aggregate(
                total_checkpoints=Count('id'),
                cercas_utilizadas=Count('cerca', distinct=True)
            )
            
            # Buscar dados de infrações
            infracoes_stats = Infrações.objects.filter(unidade=unidade).aggregate(
                total_infracoes=Count('id'),
                velocidade_media_infracoes=Avg('velocidade'),
                velocidade_maxima_infracoes=Max('velocidade')
            )
            
            dados.append({
                'ID_Unidade': unidade.id,
                'Nome': unidade.nm or '',
                'Classe': unidade.cls or '',
                'Marca': unidade.marca or '',
                'Placa': unidade.placa or '',
                'Descrição': unidade.descricao or '',
                'Total_Viagens': stats['total_viagens'] or 0,
                'Total_Quilometragem_km': float(stats['total_quilometragem'] or 0),
                'Total_Consumo_L': stats['total_consumo'] or 0,
                'Eficiencia_Media_km_L': float(stats['eficiencia_media'] or 0),
                'Velocidade_Media_km_h': stats['velocidade_media'] or 0,
                'RPM_Medio': stats['rpm_medio'] or 0,
                'Temperatura_Media_C': stats['temperatura_media'] or 0,
                'Total_Emissoes_CO2': stats['total_emissoes'] or 0,
                'Melhor_Eficiencia_km_L': float(stats['melhor_eficiencia'] or 0),
                'Pior_Eficiencia_km_L': float(stats['pior_eficiencia'] or 0),
                'Ultimo_Periodo': stats['ultimo_periodo'] or '',
                'Total_CheckPoints': checkpoints_stats['total_checkpoints'] or 0,
                'Cercas_Utilizadas': checkpoints_stats['cercas_utilizadas'] or 0,
                'Total_Infracoes': infracoes_stats['total_infracoes'] or 0,
                'Velocidade_Media_Infracoes_km_h': infracoes_stats['velocidade_media_infracoes'] or 0,
                'Velocidade_Maxima_Infracoes_km_h': infracoes_stats['velocidade_maxima_infracoes'] or 0,
            })
        
        return dados

    def coletar_dados_viagens(self, empresa):
        """Coleta dados das viagens por período da empresa"""
        # Buscar todas as viagens da empresa com filtros de eficiência
        viagens = Viagem_Base.objects.filter(
            unidade__empresa=empresa,
            Quilometragem_média__gte=1.0,
            Quilometragem_média__lte=4.0
        ).select_related('unidade').order_by('-período', 'unidade__cls', 'unidade__id')
        
        dados = []
        
        for viagem in viagens:
            dados.append({
                'Período': viagem.período or '',
                'ID_Unidade': viagem.unidade.id,
                'Nome_Unidade': viagem.unidade.nm or '',
                'Classe_Unidade': viagem.unidade.cls or '',
                'Marca_Unidade': viagem.unidade.marca or '',
                'Placa_Unidade': viagem.unidade.placa or '',
                'Quilometragem_km': float(viagem.quilometragem or 0),
                'Consumido_L': viagem.Consumido or 0,
                'Quilometragem_Media_km_L': float(viagem.Quilometragem_média or 0),
                'Horas_Motor': viagem.Horas_de_motor or '',
                'Velocidade_Media_km_h': viagem.Velocidade_média or 0,
                'RPM_Medio': viagem.RPM_médio or 0,
                'Temperatura_Media_C': viagem.Temperatura_média or 0,
                'Emissoes_CO2': viagem.Emissões_CO2 or 0,
            })
        
        return dados

    def formatar_planilhas(self, writer):
        """Aplica formatação básica nas planilhas"""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Formatação para a planilha de Unidades
        worksheet_unidades = writer.sheets['Unidades']
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Aplicar formatação no cabeçalho
        for cell in worksheet_unidades[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Ajustar largura das colunas
        for column in worksheet_unidades.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet_unidades.column_dimensions[column_letter].width = adjusted_width
        
        # Formatação para a planilha de Viagens
        worksheet_viagens = writer.sheets['Viagens por Período']
        
        # Aplicar formatação no cabeçalho
        for cell in worksheet_viagens[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Ajustar largura das colunas
        for column in worksheet_viagens.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet_viagens.column_dimensions[column_letter].width = adjusted_width