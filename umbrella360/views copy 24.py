from django.db.models import F, Avg, Sum, Q, Count, Max, Min
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.urls import reverse
from django.views import generic
from django.utils import timezone
from django.core.paginator import Paginator
from datetime import datetime
import io
import os
import json
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from .models import (
    Motorista, Caminhao, Viagem_MOT, Viagem_CAM,
    Empresa, Unidade, Viagem_Base, CheckPoint, Veiculo, Viagem_Detalhada
)
from .config import Config, ConfiguracaoManager
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go

# ========== FUNÇÕES AUXILIARES ==========

def get_empresas_disponiveis():
    """Retorna lista de empresas disponíveis no sistema"""
    return Empresa.objects.all().order_by('nome')

def get_marcas_por_empresa(empresa_id=None):
    """Retorna lista de marcas disponíveis para uma empresa específica"""
    if empresa_id and empresa_id != 'todas':
        return Unidade.objects.filter(empresa_id=empresa_id).values_list('marca', flat=True).distinct().order_by('marca')
    return Unidade.objects.values_list('marca', flat=True).distinct().order_by('marca')

def get_periodos_disponiveis():
    """Retorna lista de períodos disponíveis nos dados unificados"""
    return Viagem_Base.objects.values_list('período', flat=True).distinct().order_by('período')

def get_meses_disponiveis():
    """Retorna lista de meses disponíveis nos dados"""
    meses_mot = Viagem_MOT.objects.values_list('mês', flat=True).distinct()
    meses_cam = Viagem_CAM.objects.values_list('mês', flat=True).distinct()
    todos_meses = set(meses_mot) | set(meses_cam)
    return sorted([m for m in todos_meses if m])

# ========== FILTROS ==========

def aplicar_filtro_empresa(queryset, empresa_selecionada):
    """Aplica filtro de empresa no queryset de Viagem_Base"""
    if empresa_selecionada and empresa_selecionada != 'todas':
        return queryset.filter(unidade__empresa_id=empresa_selecionada)
    return queryset

def aplicar_filtro_marca(queryset, marca_selecionada):
    """Aplica filtro de marca no queryset de Viagem_Base"""
    if marca_selecionada and marca_selecionada != 'todas':
        return queryset.filter(unidade__marca=marca_selecionada)
    return queryset

def aplicar_filtro_periodo(queryset, periodo_selecionado):
    """Aplica filtro de período no queryset de Viagem_Base"""
    if periodo_selecionado and periodo_selecionado != 'todos':
        return queryset.filter(período=periodo_selecionado)
    return queryset

def aplicar_filtro_combustivel(queryset, filtro_combustivel):
    """Aplica filtro de combustível baseado no tipo selecionado usando configuração dinâmica"""
    if filtro_combustivel == 'sem_zero':
        return queryset.filter(Consumido__gt=0)
    elif filtro_combustivel == 'normais':
        consumo_max = Config.consumo_maximo_normal()
        return queryset.filter(Consumido__gt=0, Consumido__lte=consumo_max)
    elif filtro_combustivel == 'erros':
        consumo_max = Config.consumo_maximo_normal()
        return queryset.filter(Consumido__gt=consumo_max)
    return queryset

def aplicar_filtro_classe_unidade(queryset, classe_selecionada):
    """Aplica filtro de classe de unidade no queryset de Viagem_Base"""
    if classe_selecionada and classe_selecionada != 'todas':
        if classe_selecionada == 'veiculo':
            return queryset.filter(unidade__cls__icontains='veículo')
        elif classe_selecionada == 'motorista':
            return queryset.filter(unidade__cls__icontains='motorista')
    return queryset

def aplicar_filtro_eficiencia_minima(queryset, eficiencia_minima=1.0):
    """Aplica filtro de eficiência mínima para remover dados irreais/com erros"""
    return queryset.filter(Quilometragem_média__gte=eficiencia_minima)

def aplicar_filtro_eficiencia_maxima(queryset, eficiencia_maxima=4.0):
    """Aplica filtro de eficiência máxima para remover dados irreais/com erros"""
    return queryset.filter(Quilometragem_média__lte=eficiencia_maxima)

def aplicar_filtro_mes(queryset, mes_selecionado):
    """Aplica filtro de mês no queryset (para sistema antigo)"""
    if mes_selecionado and mes_selecionado != 'todos':
        return queryset.filter(mês=mes_selecionado)
    return queryset

def aplicar_filtros_combinados_novo(queryset, empresa_selecionada, marca_selecionada, periodo_selecionado, filtro_combustivel, classe_selecionada=None):
    """Aplica todos os filtros no queryset de Viagem_Base"""
    queryset = aplicar_filtro_empresa(queryset, empresa_selecionada)
    queryset = aplicar_filtro_marca(queryset, marca_selecionada)
    queryset = aplicar_filtro_periodo(queryset, periodo_selecionado)
    queryset = aplicar_filtro_combustivel(queryset, filtro_combustivel)
    if classe_selecionada:
        queryset = aplicar_filtro_classe_unidade(queryset, classe_selecionada)
    return queryset

def aplicar_filtros_combinados_antigo(queryset, mes_selecionado, filtro_combustivel):
    """Aplica filtros de mês e combustível no queryset (sistema antigo)"""
    queryset = aplicar_filtro_mes(queryset, mes_selecionado)
    queryset = aplicar_filtro_combustivel(queryset, filtro_combustivel)
    return queryset

# ========== PROCESSAMENTO DE REQUESTS ==========

def processar_filtros_request_novo(request):
    """Processa filtros da requisição para o novo sistema"""
    empresa_selecionada = request.GET.get('empresa', 'todas')
    marca_selecionada = request.GET.get('marca', 'todas')
    periodo_selecionado = request.GET.get('periodo', 'todos')
    filtro_combustivel = request.GET.get('filtro_combustivel', 'todos')
    classe_selecionada = request.GET.get('classe_unidade', 'todas')
    
    # Manter compatibilidade com parâmetro antigo
    remover_zero = request.GET.get('remover_zero', 'nao')
    if remover_zero == 'sim' and filtro_combustivel == 'todos':
        filtro_combustivel = 'sem_zero'
    
    return empresa_selecionada, marca_selecionada, periodo_selecionado, filtro_combustivel, classe_selecionada, remover_zero

def processar_filtros_request_antigo(request):
    """Processa e normaliza filtros da requisição (sistema antigo)"""
    mes_selecionado = request.GET.get('mes', 'todos')
    filtro_combustivel = request.GET.get('filtro_combustivel', 'todos')
    
    # Manter compatibilidade com parâmetro antigo
    remover_zero = request.GET.get('remover_zero', 'nao')
    if remover_zero == 'sim' and filtro_combustivel == 'todos':
        filtro_combustivel = 'sem_zero'
    
    return mes_selecionado, filtro_combustivel, remover_zero

# ========== CONTEXTOS ==========

def get_base_context_novo(empresa_selecionada, marca_selecionada, periodo_selecionado, filtro_combustivel, classe_selecionada, remover_zero):
    """Retorna contexto base para o novo sistema"""
    return {
        'empresa_selecionada': empresa_selecionada,
        'empresas_disponiveis': get_empresas_disponiveis(),
        'marca_selecionada': marca_selecionada,
        'marcas_disponiveis': get_marcas_por_empresa(empresa_selecionada),
        'periodo_selecionado': periodo_selecionado,
        'periodos_disponiveis': get_periodos_disponiveis(),
        'filtro_combustivel': filtro_combustivel,
        'classe_selecionada': classe_selecionada,
        'remover_zero': remover_zero,
    }

def get_base_context_antigo(mes_selecionado, filtro_combustivel, remover_zero):
    """Retorna contexto base comum para sistema antigo"""
    return {
        'mes_selecionado': mes_selecionado,
        'meses_disponiveis': get_meses_disponiveis(),
        'filtro_combustivel': filtro_combustivel,
        'remover_zero': remover_zero,
    }

# ========== CÁLCULOS E ESTATÍSTICAS ==========

def calcular_stats_marca(viagens_filtradas, marca):
    """Calcula estatísticas agregadas para uma marca específica (sistema antigo)"""
    return viagens_filtradas.filter(agrupamento__marca=marca).aggregate(
        total_quilometragem=Sum('quilometragem'),
        total_consumido=Sum('Consumido'),
        media_quilometragem=Avg('Quilometragem_média'),
        media_velocidade=Avg('Velocidade_média'),
        media_rpm=Avg('RPM_médio'),
        media_temperatura=Avg('Temperatura_média'),
        total_emissoes=Sum('Emissões_CO2')
    )

def calcular_stats_marca_novo(viagens_filtradas, marca):
    """Calcula estatísticas agregadas para uma marca específica usando dados unificados"""
    viagens_marca = viagens_filtradas.filter(unidade__marca=marca)
    
    # Estatísticas básicas
    stats = viagens_marca.aggregate(
        total_quilometragem=Sum('quilometragem'),
        total_consumido=Sum('Consumido'),
        media_quilometragem=Avg('Quilometragem_média'),
        media_velocidade=Avg('Velocidade_média'),
        media_rpm=Avg('RPM_médio'),
        media_temperatura=Avg('Temperatura_média'),
        total_emissoes=Sum('Emissões_CO2'),
        total_veiculos=Count('unidade', distinct=True)
    )
    
    # Calcular médias por veículo
    total_veiculos = stats['total_veiculos'] or 1  # Evitar divisão por zero
    
    stats['quilometragem_por_veiculo'] = (stats['total_quilometragem'] or 0) / total_veiculos
    stats['consumo_por_veiculo'] = (stats['total_consumido'] or 0) / total_veiculos
    stats['emissoes_por_veiculo'] = (stats['total_emissoes'] or 0) / total_veiculos
    # A eficiência média já é uma média, não precisa dividir por veículos
    stats['eficiencia_por_veiculo'] = stats['media_quilometragem'] or 0
    
    return stats

def calcular_stats_empresa(viagens_filtradas, empresa_nome):
    """Calcula estatísticas agregadas para uma empresa específica"""
    return viagens_filtradas.filter(unidade__empresa__nome=empresa_nome).aggregate(
        total_quilometragem=Sum('quilometragem'),
        total_consumido=Sum('Consumido'),
        media_quilometragem=Avg('Quilometragem_média'),
        media_velocidade=Avg('Velocidade_média'),
        media_rpm=Avg('RPM_médio'),
        media_temperatura=Avg('Temperatura_média'),
        total_emissoes=Sum('Emissões_CO2'),
        total_unidades=Count('unidade', distinct=True)
    )

def obter_unidades_com_stats(empresa_selecionada=None):
    """Obtém todas as unidades com suas estatísticas básicas"""
    unidades_query = Unidade.objects.select_related('empresa').all()
    
    if empresa_selecionada and empresa_selecionada != 'todas':
        unidades_query = unidades_query.filter(empresa_id=empresa_selecionada)
    
    unidades_list = []
    
    for unidade in unidades_query.order_by('empresa__nome', 'cls', 'id'):
        # Buscar estatísticas da unidade (filtrar eficiência abaixo de 4 km/L)
        viagens_unidade = Viagem_Base.objects.filter(unidade=unidade)

        
        stats = viagens_unidade.aggregate(
            total_viagens=Count('id'),
            eficiencia_media=Avg('Quilometragem_média'),
            ultima_viagem_periodo=Max('período')
        )
        
        # Buscar a última viagem completa
        ultima_viagem = viagens_unidade.order_by('-período').first()
        
        # Adicionar dados calculados à unidade
        unidade.total_viagens = stats['total_viagens'] or 0
        unidade.eficiencia_media = stats['eficiencia_media']
        unidade.ultima_viagem = ultima_viagem
        
        unidades_list.append(unidade)
    
    return unidades_list

# ========== VIEWS PRINCIPAIS ==========

def index(request):
    """View principal do dashboard"""
    mes_selecionado, filtro_combustivel, remover_zero = processar_filtros_request_antigo(request)
    context = get_base_context_antigo(mes_selecionado, filtro_combustivel, remover_zero)
    return render(request, "umbrella360/index.html", context)

def report_novo(request):
    """Nova view para relatórios usando dados unificados com suporte a múltiplas empresas"""
    # Obter filtros
    empresa_selecionada, marca_selecionada, periodo_selecionado, filtro_combustivel, classe_selecionada, remover_zero = processar_filtros_request_novo(request)
    
    # Contadores gerais
    total_empresas = Empresa.objects.count()
    total_unidades = Unidade.objects.count()
    
    # Aplicar filtros nas viagens da tabela unificada
    viagens_base = Viagem_Base.objects.select_related('unidade', 'unidade__empresa')
    viagens_filtradas = aplicar_filtros_combinados_novo(
        viagens_base, empresa_selecionada, marca_selecionada, 
        periodo_selecionado, filtro_combustivel, classe_selecionada
    )

    #filtra as viagens pelos ultimos 30 dias
    #viagens_filtradas = viagens_filtradas.filter(período="Últimos 30 dias")
    
    # Filtrar viagens com eficiência abaixo de 4 km/L (dados irreais/com erros)
    viagens_filtradas = aplicar_filtro_eficiencia_minima(viagens_filtradas)
    viagens_filtradas = aplicar_filtro_eficiencia_maxima(viagens_filtradas)
    
    # Separar por tipo de unidade (motoristas e veículos)
    viagens_motoristas = viagens_filtradas.filter(unidade__cls__icontains='motorista').order_by('-Quilometragem_média')[:10]
    viagens_veiculos = viagens_filtradas.filter(unidade__cls__icontains='veículo').order_by('-Quilometragem_média')[:10]
    
    # Cálculos agregados
    total_quilometragem = viagens_filtradas.aggregate(total=Sum('quilometragem'))['total'] or 0
    total_emissoes = viagens_filtradas.aggregate(total=Sum('Emissões_CO2'))['total'] or 0
    rpm_medio = viagens_filtradas.aggregate(media=Avg('RPM_médio'))['media'] or 0
    velocidade_media = viagens_filtradas.aggregate(media=Avg('Velocidade_média'))['media'] or 0
    media_quilometragem = viagens_filtradas.aggregate(media=Avg('Quilometragem_média'))['media'] or 0
    
    # Consumo com limite configurável
    consumo_max_normal = Config.consumo_maximo_normal()
    total_consumo = viagens_filtradas.filter(Consumido__lte=consumo_max_normal).aggregate(total=Sum('Consumido'))['total'] or 0
    
    # Calculadora de custos
    custo_diesel = Config.custo_diesel()
    
    # Calcular média real de eficiência dos dados filtrados
    media_km_atual_calculada = viagens_filtradas.filter(
        Consumido__gt=0, Consumido__lte=consumo_max_normal
    ).aggregate(media=Avg('Quilometragem_média'))['media']
    
    media_km_atual = float(media_km_atual_calculada) if media_km_atual_calculada and media_km_atual_calculada > 0 else Config.media_km_atual()
    media_km_objetivo = Config.media_km_objetivo()
    
    # Cálculos de economia
    km_atual_litros = float(total_quilometragem) / media_km_atual if media_km_atual > 0 else 0
    km_objetivo_litros = float(total_quilometragem) / media_km_objetivo if media_km_objetivo > 0 else 0
    
    custo_atual = km_atual_litros * custo_diesel
    custo_objetivo = km_objetivo_litros * custo_diesel
    economia_potencial = custo_atual - custo_objetivo
    
    # Estatísticas por marca (dinâmicas baseadas nos dados filtrados)
    marcas_stats = {}
    marcas_disponiveis = viagens_filtradas.values_list('unidade__marca', flat=True).distinct()
    
    for marca in marcas_disponiveis:
        if marca:  # Evitar valores None
            stats = calcular_stats_marca_novo(viagens_filtradas, marca)
            stats['marca'] = marca
            marcas_stats[marca] = stats
    
    # Estatísticas por empresa
    empresas_stats = {}
    if empresa_selecionada == 'todas':
        empresas_disponiveis = viagens_filtradas.values_list('unidade__empresa__nome', flat=True).distinct()
        for empresa_nome in empresas_disponiveis:
            if empresa_nome:
                stats = calcular_stats_empresa(viagens_filtradas, empresa_nome)
                empresas_stats[empresa_nome] = stats
    
    # Aplicar filtros na busca de unidades
    unidades_filtradas = Unidade.objects.select_related('empresa').all()
    if empresa_selecionada != 'todas':
        unidades_filtradas = unidades_filtradas.filter(empresa_id=empresa_selecionada)
    if classe_selecionada != 'todas':
        if classe_selecionada == 'veiculo':
            unidades_filtradas = unidades_filtradas.filter(cls__icontains='veículo')
        elif classe_selecionada == 'motorista':
            unidades_filtradas = unidades_filtradas.filter(cls__icontains='motorista')
    
    todas_unidades = []
    for unidade in unidades_filtradas.order_by('empresa__nome', 'cls', 'id'):
        viagens_unidade_filtradas = viagens_filtradas.filter(unidade=unidade)
        stats = viagens_unidade_filtradas.aggregate(
            total_viagens=Count('id'),
            eficiencia_media=Avg('Quilometragem_média'),
            ultima_viagem_periodo=Max('período')
        )
        unidade.total_viagens = stats['total_viagens'] or 0
        unidade.eficiencia_media = stats['eficiencia_media']
        unidade.periodo_stats = stats['ultima_viagem_periodo']
        todas_unidades.append(unidade)
    
    context = get_base_context_novo(empresa_selecionada, marca_selecionada, periodo_selecionado, filtro_combustivel, classe_selecionada, remover_zero)
    context.update({
        'viagens_motoristas': viagens_motoristas,
        'viagens_veiculos': viagens_veiculos,
        'marcas_stats': marcas_stats,
        'empresas_stats': empresas_stats,
        'todas_unidades': todas_unidades,
        'total_empresas': total_empresas,
        'total_unidades': total_unidades,
        'total_quilometragem': total_quilometragem,
        'total_consumo': total_consumo,
        'total_emissoes': total_emissoes,
        'rpm_medio': rpm_medio,
        'velocidade_media': velocidade_media,
        'custo_diesel': custo_diesel,
        'media_km_objetivo': media_km_objetivo,
        'media_km_atual': media_km_atual,
        'custo_atual': custo_atual,
        'custo_objetivo': custo_objetivo,
        'economia_potencial': economia_potencial,
        'media_quilometragem': media_quilometragem,
    })
    
    return render(request, 'umbrella360/report_novo.html', context)

def lista_unidades(request):
    """View para listar unidades do sistema - filtra por empresa se logado"""
    # Verificar se há empresa logada
    empresa_logada_id = request.session.get('empresa_logada')
    empresa_logada = None
    
    if empresa_logada_id:
        try:
            empresa_logada = Empresa.objects.get(id=empresa_logada_id)
            # Se logado, mostrar apenas unidades da empresa
            todas_unidades = obter_unidades_com_stats(empresa_logada_id)
            total_empresas = 1  # Sempre 1 quando filtrado por empresa
        except Empresa.DoesNotExist:
            # Se empresa não existe mais, limpar sessão e mostrar todas
            request.session.flush()
            todas_unidades = obter_unidades_com_stats(None)
            total_empresas = Empresa.objects.count()
    else:
        # Se não logado, mostrar todas as unidades
        todas_unidades = obter_unidades_com_stats(None)
        total_empresas = Empresa.objects.count()
    
    # Contadores básicos
    total_unidades = len(todas_unidades)
    motoristas_count = len([u for u in todas_unidades if u.cls == 'Motorista'])
    veiculos_count = len([u for u in todas_unidades if u.cls == 'Veículo'])

    context = {
        'todas_unidades': todas_unidades,
        'total_unidades': total_unidades,
        'total_empresas': total_empresas,
        'motoristas_count': motoristas_count,
        'veiculos_count': veiculos_count,
        'empresa_logada': empresa_logada,  # Passar info da empresa logada para template
    }
    
    return render(request, 'umbrella360/lista_unidades.html', context)

def detalhes_unidade(request, unidade_id):
    """View para mostrar detalhes completos de uma unidade específica"""
    unidade = get_object_or_404(Unidade, id=unidade_id)
    
    # 🔥 NOVO: Obter filtros de data
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    periodo_dias = request.GET.get('periodo_dias', '1')  # Padrão: dia atual
    
    
    # Obter todas as viagens da unidade
    viagens = Viagem_Base.objects.filter(unidade=unidade).order_by('-período')

    # Calcular estatísticas gerais
    stats_gerais = viagens.aggregate(
        total_viagens=Count('id'),
        total_quilometragem=Sum('quilometragem'),
        total_consumo=Sum('Consumido'),
        media_eficiencia=Avg('Quilometragem_média'),
        media_velocidade=Avg('Velocidade_média'),
        media_rpm=Avg('RPM_médio'),
        media_temperatura=Avg('Temperatura_média'),
        total_emissoes=Sum('Emissões_CO2'),
        melhor_eficiencia=Max('Quilometragem_média'),
        pior_eficiencia=Min('Quilometragem_média')
    )
    #estatisticas filtradas
    viagens = Viagem_Base.objects.filter(unidade=unidade).order_by('-período')
    #filtrar viagens pelo período dos "Últimos 30 dias"
    viagens_filtradas = viagens.filter(período="Últimos 30 dias")
    stats_filtrados = viagens_filtradas.aggregate(
        total_viagens=Count('id'),
        total_quilometragem=Sum('quilometragem'),
        total_consumo=Sum('Consumido'),
        media_eficiencia=Avg('Quilometragem_média'),
        media_velocidade=Avg('Velocidade_média'),
        media_rpm=Avg('RPM_médio'),
        media_temperatura=Avg('Temperatura_média'),
        total_emissoes=Sum('Emissões_CO2'),
        melhor_eficiencia=Max('Quilometragem_média'),
        pior_eficiencia=Min('Quilometragem_média')
    )

    # GRÁFICO TIMELINE RPM vs TEMPO
    # Buscar dados do Viagem_eco para a unidade
    from .models import Viagem_eco
    from datetime import datetime, timedelta
    import json
    
    # Buscar dados dos últimos 7 dias ou últimos 1000 registros (o que for menor)
    #data_limite = timezone.now() - timedelta(days=7)
    
    viagens_eco = Viagem_eco.objects.filter(
        unidade_id=unidade.id
    ).order_by('-timestamp')  # Limitar para performance


    
    # 🔥 NOVO: Aplicar filtros de data
    if data_inicio and data_fim:
        try:
            dt_inicio = datetime.strptime(data_inicio, '%Y-%m-%d')
            dt_fim = datetime.strptime(data_fim, '%Y-%m-%d')
            timestamp_inicio = int(dt_inicio.timestamp())
            timestamp_fim = int(dt_fim.replace(hour=23, minute=59, second=59).timestamp())
            viagens_eco = viagens_eco.filter(
                timestamp__gte=timestamp_inicio,
                timestamp__lte=timestamp_fim
            )
        except ValueError:
            pass
    elif periodo_dias and periodo_dias != 'todos':
        try:
            dias = int(periodo_dias)
            if dias == 0:  # Dia atual
                hoje = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
                timestamp_inicio = int(hoje.timestamp())
                timestamp_fim = int(datetime.now().timestamp())
            else:
                data_limite = datetime.now() - timedelta(days=dias)
                timestamp_inicio = int(data_limite.timestamp())
                timestamp_fim = int(datetime.now().timestamp())
            
            viagens_eco = viagens_eco.filter(
                timestamp__gte=timestamp_inicio,
                timestamp__lte=timestamp_fim
            )
        except ValueError:
            pass
    



    viagens_detalhadas = Viagem_Detalhada.objects.filter(
        unidade_id=unidade.id
    ).order_by('-timestamp_final')
    
    # Preparar dados para o gráfico, remove rpm acima de 3500

    
    timeline_data = []
    if viagens_eco.exists():
        for viagem_eco in reversed(viagens_eco):  # Reverter para ordem cronológica
            try:
                # Converter timestamp unix para datetime
                timestamp_int = int(viagem_eco.timestamp)
                dt = datetime.fromtimestamp(timestamp_int)


                # Filtrar valores de RPM muito altos (acima de 3500)

                if float(viagem_eco.rpm) and float(viagem_eco.rpm/8) <= 3500:
                    timeline_data.append({
                        'timestamp': dt.strftime('%Y-%m-%d %H:%M:%S'),
                        'rpm': float(viagem_eco.rpm/8) if viagem_eco.rpm else 0,
                        'velocidade': float(viagem_eco.velocidade) if hasattr(viagem_eco, 'velocidade') and viagem_eco.velocidade else 0,
                    'altitude': float(viagem_eco.altitude) if hasattr(viagem_eco, 'altitude') and viagem_eco.altitude else 0,
                    'temperatura': float(viagem_eco.temperatura_motor) if hasattr(viagem_eco, 'temperatura_motor') and viagem_eco.temperatura_motor else 0,
                    'odometro': float(viagem_eco.odometro) if hasattr(viagem_eco, 'odometro') and viagem_eco.odometro else 0,
                    'consumo': None,  # Não disponível em Viagem_eco
                    'source': 'eco'
                })
            except (ValueError, TypeError, OSError):
                continue  # Pular timestamps inválidos

            # Depois, processar dados do Viagem_Detalhada (dados de consumo)
    consumo_data = []
    if viagens_detalhadas.exists():
        for viagem_det in viagens_detalhadas:
            try:
                if viagem_det.timestamp_inicial:
                    # Converter timestamp inicial para datetime
                    dt_inicial = datetime.fromtimestamp(viagem_det.timestamp_inicial)
                    
                    # Adicionar ponto inicial da viagem
                    consumo_data.append({
                        'timestamp': dt_inicial.strftime('%Y-%m-%d %H:%M:%S'),
                        'consumo': 0,  # Início da viagem, consumo zero
                        'consumo_acumulado': 0,
                        'eficiencia': float(viagem_det.Quilometragem_média) if viagem_det.Quilometragem_média else 0,
                        'source': 'detalhada'
                    })
                    
                    # Se houver timestamp final, adicionar ponto final
                    if viagem_det.timestamp_final:
                        dt_final = datetime.fromtimestamp(viagem_det.timestamp_final)
                        
                        consumo_data.append({
                            'timestamp': dt_final.strftime('%Y-%m-%d %H:%M:%S'),
                            'consumo': float(viagem_det.Consumido) if viagem_det.Consumido else 0,
                            'consumo_acumulado': float(viagem_det.Consumido) if viagem_det.Consumido else 0,
                            'eficiencia': float(viagem_det.Quilometragem_média) if viagem_det.Quilometragem_média else 0,
                            'source': 'detalhada'
                        })
                    else:
                        # Se não houver timestamp final, usar o inicial + duração estimada
                        # ou apenas marcar como ponto único
                        consumo_data[-1]['consumo'] = float(viagem_det.Consumido) if viagem_det.Consumido else 0
                        consumo_data[-1]['consumo_acumulado'] = float(viagem_det.Consumido) if viagem_det.Consumido else 0
                        
            except (ValueError, TypeError, OSError):
                continue  # Pular timestamps inválidos
    
    # Combinar os dados (se necessário) ou manter separados
    # Para o gráfico, vamos manter os dados de timeline_data para RPM/velocidade/altitude
    # E criar um dataset separado para consumo
    
    # Converter dados para JSON para uso no template
    timeline_data_json = json.dumps(timeline_data)
    consumo_data_json = json.dumps(consumo_data)
    
    # Estatísticas do gráfico
    timeline_stats = {
        'total_pontos': len(timeline_data),
        'rpm_medio': sum(d['rpm'] for d in timeline_data) / len(timeline_data) if timeline_data else 0,
        'rpm_maximo': max(d['rpm'] for d in timeline_data) if timeline_data else 0,
        'rpm_minimo': min(d['rpm'] for d in timeline_data) if timeline_data else 0,
        'velocidade_media': sum(d['velocidade'] for d in timeline_data) / len(timeline_data) if timeline_data else 0,
        'periodo_inicio': timeline_data[0]['timestamp'] if timeline_data else None,
        'periodo_fim': timeline_data[-1]['timestamp'] if timeline_data else None,
        #faixas de RPM, excluidos valores abaixo de 500 rpm
        'faixas_rpm': {
            'azul': len([d for d in timeline_data if 350 <= d['rpm'] < 799]),
            'verde': len([d for d in timeline_data if 800 <= d['rpm'] < 1300]),
            'amarela': len([d for d in timeline_data if 1301 <= d['rpm'] < 2300]),
            'vermelha': len([d for d in timeline_data if d['rpm'] >= 2301]),
        },
        'total_viagens_consumo': len(consumo_data),
        'consumo_total': sum(d['consumo'] for d in consumo_data) if consumo_data else 0,
        'eficiencia_media': sum(d['eficiencia'] for d in consumo_data) / len(consumo_data) if consumo_data else 0,
    }

    # Estatísticas por período
    stats_por_periodo = viagens.values('período').annotate(
        viagens_periodo=Count('id'),
        quilometragem_periodo=Sum('quilometragem'),
        consumo_periodo=Sum('Consumido'),
        eficiencia_periodo=Avg('Quilometragem_média')
    ).order_by('-período')
    
    # Calcular custos
    custo_diesel = Config.custo_diesel()
    custo_total = (stats_gerais['total_consumo'] or 0) * custo_diesel
    
    # Comparar com médias do sistema
    media_sistema = Viagem_Base.objects.filter(
        unidade__cls=unidade.cls
    ).aggregate(
        media_eficiencia_sistema=Avg('Quilometragem_média')
    )['media_eficiencia_sistema'] or 0
    
    # Dados de CheckPoints
    checkpoints = CheckPoint.objects.filter(unidade=unidade).order_by('-data_entrada')
    
    # Estatísticas de checkpoints
    total_checkpoints_unidade = checkpoints.count()
    cercas_utilizadas_unidade = checkpoints.values('cerca').distinct().count()
    
    # Duração média dos checkpoints (convertendo para minutos)
    duracao_media_checkpoints = 0
    if checkpoints.exists():
        # Calcular duração média em minutos
        checkpoints_com_duracao = checkpoints.exclude(duracao__isnull=True)
        if checkpoints_com_duracao.exists():
            duracao_total_segundos = sum([
                cp.duracao.total_seconds() 
                for cp in checkpoints_com_duracao 
                if cp.duracao
            ])
            duracao_media_checkpoints = int(duracao_total_segundos / checkpoints_com_duracao.count() / 60)
    
    # Cercas mais utilizadas pela unidade
    cercas_unidade = checkpoints.values('cerca').annotate(
        total_passagens=Count('id')
    ).order_by('-total_passagens')[:10]
    
    # Últimas passagens
    checkpoints_recentes_unidade = checkpoints[:10]
    
    # Histórico detalhado de checkpoints
    checkpoints_detalhados_unidade = checkpoints[:20]
    
    # DADOS DAS INFRAÇÕES
    # Buscar infrações da unidade
    from .models import Infrações
    infracoes = Infrações.objects.filter(unidade=unidade).order_by('-data')
    
    # Estatísticas de infrações
    total_infracoes_unidade = infracoes.count()
    
    # Velocidade média das infrações desta unidade
    stats_infracoes_unidade = infracoes.aggregate(
        velocidade_media=Avg('velocidade'),
        velocidade_maxima=Max('velocidade'),
        limite_medio=Avg('limite'),
        limite_mais_infringido=Max('limite')
    )
    
    # Limites mais infringidos por esta unidade
    limites_unidade = infracoes.values('limite').annotate(
        total_ocorrencias=Count('id')
    ).order_by('-total_ocorrencias')[:5]
    
    # Últimas infrações
    infracoes_recentes_unidade = infracoes[:10]
    
    # Histórico detalhado de infrações
    infracoes_detalhadas_unidade = infracoes[:20]
    
    context = {
        'unidade': unidade,
        'viagens': viagens[:20],  # Últimas 20 viagens
        'stats_gerais': stats_gerais,
        'stats_filtrados': stats_filtrados,
        'stats_por_periodo': stats_por_periodo,
        'custo_total': custo_total,
        'custo_diesel': custo_diesel,
        'media_sistema': media_sistema,
        'total_viagens': viagens.count(),
        # Dados de checkpoints
        'total_checkpoints_unidade': total_checkpoints_unidade,
        'cercas_utilizadas_unidade': cercas_utilizadas_unidade,
        'duracao_media_checkpoints': duracao_media_checkpoints,
        'cercas_unidade': cercas_unidade,
        'checkpoints_recentes_unidade': checkpoints_recentes_unidade,
        'checkpoints_detalhados_unidade': checkpoints_detalhados_unidade,
        # Dados de infrações
        'total_infracoes_unidade': total_infracoes_unidade,
        'velocidade_media_infracoes_unidade': stats_infracoes_unidade['velocidade_media'] or 0,
        'velocidade_maxima_infracoes_unidade': stats_infracoes_unidade['velocidade_maxima'] or 0,
        'limite_medio_infracoes_unidade': stats_infracoes_unidade['limite_medio'] or 0,
        'limites_unidade': limites_unidade,
        'infracoes_recentes_unidade': infracoes_recentes_unidade,
        'infracoes_detalhadas_unidade': infracoes_detalhadas_unidade,
        #
        'odometro_unidade': unidade.odometro,
        # Dados do gráfico timeline
        'timeline_data_json': timeline_data_json,
        'timeline_stats': timeline_stats,
        'timeline_data_json': timeline_data_json,
        'consumo_data_json': consumo_data_json,  # Novo
                # 🔥 NOVO: Passar filtros para o template
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'periodo_dias': periodo_dias,
    }
    

    
    return render(request, 'umbrella360/detalhes_unidade.html', context)

# ========== VIEWS LEGADAS (Sistema Antigo) ==========

def report(request):
    """View para relatórios do sistema antigo (mantida para compatibilidade)"""
    mes_selecionado, filtro_combustivel, remover_zero = processar_filtros_request_antigo(request)
    
    total_motoristas = Motorista.objects.count()
    total_caminhoes = Caminhao.objects.count()

    # Aplicar filtros nas viagens
    viagens_motoristas_base = Viagem_MOT.objects.select_related('agrupamento')
    viagens_caminhoes_base = Viagem_CAM.objects.select_related('agrupamento')
    
    viagens_motoristas_filtradas = aplicar_filtros_combinados_antigo(viagens_motoristas_base, mes_selecionado, filtro_combustivel)
    viagens_caminhoes_filtradas = aplicar_filtros_combinados_antigo(viagens_caminhoes_base, mes_selecionado, filtro_combustivel)

    # Get top 5 viagens de motoristas e caminhões ordenadas por média de consumo
    viagens_motoristas = viagens_motoristas_filtradas.order_by('-Quilometragem_média')[:5]
    viagens_caminhoes = viagens_caminhoes_filtradas.order_by('-Quilometragem_média')[:5]

    # Cálculos com base nas viagens de caminhões filtradas
    total_quilometragem_caminhoes = viagens_caminhoes_filtradas.aggregate(total=Sum('quilometragem'))['total']
    total_emissoes_caminhoes = viagens_caminhoes_filtradas.aggregate(total=Sum('Emissões_CO2'))['total']
    rpm_medio_caminhoes = viagens_caminhoes_filtradas.aggregate(media=Avg('RPM_médio'))['media']
    velocidade_media_caminhoes = viagens_caminhoes_filtradas.aggregate(media=Avg('Velocidade_média'))['media']

    # Usar limite de consumo configurável
    consumo_max_normal = Config.consumo_maximo_normal()
    total_consumo_caminhoes = viagens_caminhoes_filtradas.filter(Consumido__lte=consumo_max_normal).aggregate(total=Sum('Consumido'))['total']

    # Calculadora Monetária
    custo_diesel = Config.custo_diesel()
    
    # Calcular média real de quilometragem por litro dos dados filtrados
    media_km_atual_calculada = viagens_caminhoes_filtradas.filter(Consumido__gt=0, Consumido__lte=consumo_max_normal).aggregate(
        media=Avg('Quilometragem_média')
    )['media']
    
    # Usar média calculada dinamicamente, ou obter da configuração como fallback
    media_km_atual = float(media_km_atual_calculada) if media_km_atual_calculada and media_km_atual_calculada > 0 else Config.media_km_atual()
    media_km_fixa = Config.media_km_objetivo()
    
    # Cálculos de economia
    km = float(total_quilometragem_caminhoes) if total_quilometragem_caminhoes is not None else 0.0
    km_objetivo = km / media_km_fixa if media_km_fixa > 0 else 0.0
    km_atual = km / media_km_atual if media_km_atual > 0 else 0.0
    
    custo_objetivo = km_objetivo * custo_diesel
    custo_atual = km_atual * custo_diesel
    economia_potencial = custo_atual - custo_objetivo

    # Get aggregated data by brand (através das viagens filtradas)
    scania_stats = calcular_stats_marca(viagens_caminhoes_filtradas, 'Scania')
    volvo_stats = calcular_stats_marca(viagens_caminhoes_filtradas, 'Volvo')
    
    # Add brand name to the stats dictionaries
    scania_stats['marca'] = 'Scania'
    volvo_stats['marca'] = 'Volvo'
    
    context = get_base_context_antigo(mes_selecionado, filtro_combustivel, remover_zero)
    context.update({
        'viagens_motoristas': viagens_motoristas,
        'viagens_caminhoes': viagens_caminhoes,
        'scania_stats': scania_stats,
        'volvo_stats': volvo_stats,
        'total_motoristas': total_motoristas,
        'total_caminhoes': total_caminhoes,
        'total_quilometragem_caminhoes': total_quilometragem_caminhoes,
        'total_consumo_caminhoes': total_consumo_caminhoes,
        'custo_diesel': custo_diesel,
        'media_km_fixa': media_km_fixa,
        'media_km_atual': media_km_atual,
        'km_objetivo': km_objetivo,
        'custo_objetivo': custo_objetivo,
        'custo_atual': custo_atual,
        'km_atual': km_atual,
        'economia_potencial': economia_potencial,
        'emissoes_caminhoes': total_emissoes_caminhoes,
        'rpm_medio_caminhoes': rpm_medio_caminhoes,
        'velocidade_media_caminhoes': velocidade_media_caminhoes,
    })

    return render(request, 'umbrella360/report.html', context)

def motoristas(request):
    """View para listar motoristas (sistema antigo)"""
    mes_selecionado, filtro_combustivel, remover_zero = processar_filtros_request_antigo(request)
    
    total_motoristas = Motorista.objects.count()
    total_caminhoes = Caminhao.objects.count()
    
    # Aplicar filtros nas viagens de motoristas
    viagens_motoristas_base = Viagem_MOT.objects.select_related('agrupamento')
    viagens_motoristas = aplicar_filtros_combinados_antigo(viagens_motoristas_base, mes_selecionado, filtro_combustivel).order_by('-Quilometragem_média')
    
    context = get_base_context_antigo(mes_selecionado, filtro_combustivel, remover_zero)
    context.update({
        'viagens_motoristas': viagens_motoristas, 
        'total_motoristas': total_motoristas, 
        'total_caminhoes': total_caminhoes,
    })
    
    return render(request, 'umbrella360/motoristas.html', context)

def caminhoes(request):
    """View para listar caminhões (sistema antigo)"""
    mes_selecionado, filtro_combustivel, remover_zero = processar_filtros_request_antigo(request)
    
    total_motoristas = Motorista.objects.count()
    total_caminhoes = Caminhao.objects.count()
    
    # Aplicar filtros nas viagens de caminhões
    viagens_caminhoes_base = Viagem_CAM.objects.select_related('agrupamento')
    viagens_caminhoes_filtradas = aplicar_filtros_combinados_antigo(viagens_caminhoes_base, mes_selecionado, filtro_combustivel)
    viagens_caminhoes = viagens_caminhoes_filtradas.order_by('-Quilometragem_média')
    
    # Calculate statistics for each brand dynamically usando viagens filtradas
    scania_stats = calcular_stats_marca(viagens_caminhoes_filtradas, 'Scania')
    volvo_stats = calcular_stats_marca(viagens_caminhoes_filtradas, 'Volvo')
    
    # Count vehicles by brand
    scania_count = Caminhao.objects.filter(marca='Scania').count()
    volvo_count = Caminhao.objects.filter(marca='Volvo').count()
    
    # Add brand name and vehicle count to the stats dictionaries
    scania_stats['marca'] = 'Scania'
    scania_stats['count_veiculos'] = scania_count
    volvo_stats['marca'] = 'Volvo'
    volvo_stats['count_veiculos'] = volvo_count

    # Generate emissions chart usando dados das viagens filtradas
    dados = viagens_caminhoes_filtradas.values('agrupamento__marca', 'Emissões_CO2')
    df = pd.DataFrame(dados)
    
    if not df.empty:
        df_grouped = df.groupby('agrupamento__marca')['Emissões_CO2'].sum().reset_index()
        chart = px.pie(df_grouped, values='Emissões_CO2', names='agrupamento__marca',
                      title='Emissões de CO2 por Marca')
        chart.update_traces(textposition='inside', textinfo='percent+label')
    else:
        chart = px.pie(values=[1], names=['Sem dados'], title='Emissões de CO2 por Marca')

    context = get_base_context_antigo(mes_selecionado, filtro_combustivel, remover_zero)
    context.update({
        'viagens_caminhoes': viagens_caminhoes, 
        'total_motoristas': total_motoristas, 
        'total_caminhoes': total_caminhoes,
        'scania_stats': scania_stats,
        'volvo_stats': volvo_stats,
        'grafico': chart,
    })

    return render(request, 'umbrella360/caminhoes.html', context)

def grafico_emissoes_por_marca(request):
    """View para exibir gráficos de emissões por marca (sistema antigo)"""
    mes_selecionado, filtro_combustivel, remover_zero = processar_filtros_request_antigo(request)
    
    # Aplicar filtros nas viagens
    viagens_caminhoes_base = Viagem_CAM.objects.select_related('agrupamento')
    viagens_caminhoes_filtradas = aplicar_filtros_combinados_antigo(viagens_caminhoes_base, mes_selecionado, filtro_combustivel)
    
    # Gráficos diversos (implementação completa seria necessária aqui)
    # Por brevidade, apenas retorno um contexto básico
    
    context = get_base_context_antigo(mes_selecionado, filtro_combustivel, remover_zero)
    context.update({
        'titulo': 'Gráficos de Emissões por Marca',
    })

    return render(request, 'umbrella360/grafico_pizza.html', context)

# ========== APIs PARA FILTROS DINÂMICOS ==========

def api_marcas_todas(request):
    """API que retorna todas as marcas disponíveis"""
    marcas = list(Unidade.objects.values_list('marca', flat=True).distinct())
    marcas = [marca for marca in marcas if marca]  # Removes None values
    return JsonResponse(marcas, safe=False)

def api_marcas_por_empresa(request, empresa_id):
    """API que retorna marcas disponíveis para uma empresa específica"""
    marcas = list(Unidade.objects.filter(empresa_id=empresa_id).values_list('marca', flat=True).distinct())
    marcas = [marca for marca in marcas if marca]  # Removes None values
    return JsonResponse(marcas, safe=False)

def login_view(request):
    """View para página de login por empresa"""
    if request.method == 'POST':
        empresa_nome = request.POST.get('empresa')
        senha = request.POST.get('senha')
        
        try:
            empresa = Empresa.objects.get(nome=empresa_nome, senha=senha)
            # Salvar a empresa na sessão
            request.session['empresa_logada'] = empresa.id
            request.session['empresa_nome'] = empresa.nome
            # Redirecionar para o index
            return HttpResponseRedirect(reverse('index'))
        except Empresa.DoesNotExist:
            context = {
                'error': 'Empresa ou senha incorreta.',
                'empresas': Empresa.objects.all().order_by('nome')
            }
            return render(request, 'umbrella360/login.html', context)
    
    # GET request
    context = {
        'empresas': Empresa.objects.all().order_by('nome')
    }
    return render(request, 'umbrella360/login.html', context)

def report_empresa(request):
    """View para relatório específico da empresa logada"""
    # Verificar se há empresa logada
    empresa_logada_id = request.session.get('empresa_logada')
    if not empresa_logada_id:
        # Se não estiver logado, redirecionar para login
        return HttpResponseRedirect(reverse('login'))
    
    try:
        empresa_logada = Empresa.objects.get(id=empresa_logada_id)
    except Empresa.DoesNotExist:
        # Se empresa não existe mais, limpar sessão e redirecionar
        request.session.flush()
        return HttpResponseRedirect(reverse('login'))
    
    # Forçar filtros para a empresa logada
    empresa_selecionada = str(empresa_logada_id)
    
    # Obter outros filtros do request
    marca_selecionada = request.GET.get('marca', 'todas')
    periodo_selecionado = request.GET.get('periodo', 'todos')
    filtro_combustivel = request.GET.get('filtro_combustivel', 'todos')
    classe_selecionada = request.GET.get('classe', 'todas')
    remover_zero = request.GET.get('remover_zero') == 'on'
    
    # Contadores específicos da empresa
    total_unidades_empresa = Unidade.objects.filter(empresa_id=empresa_logada_id).count()
    
    # Aplicar filtros nas viagens da tabela unificada - APENAS DA EMPRESA LOGADA
    viagens_base = Viagem_Base.objects.select_related('unidade', 'unidade__empresa').filter(unidade__empresa_id=empresa_logada_id)
    viagens_filtradas = aplicar_filtros_combinados_novo(
        viagens_base, empresa_selecionada, marca_selecionada, 
        periodo_selecionado, filtro_combustivel, classe_selecionada
    )
    
    # Filtrar viagens com eficiência abaixo de 4 km/L (dados irreais/com erros)
    viagens_filtradas = aplicar_filtro_eficiencia_minima(viagens_filtradas)
    viagens_filtradas = aplicar_filtro_eficiencia_maxima(viagens_filtradas)
    
    # Separar por tipo de unidade (motoristas e veículos)
    viagens_motoristas = viagens_filtradas.filter(unidade__cls__icontains='motorista').order_by('-Quilometragem_média')[:10]
    viagens_veiculos = viagens_filtradas.filter(unidade__cls__icontains='veículo').order_by('-Quilometragem_média')[:10]
    
    # Cálculos agregados
    total_quilometragem = viagens_filtradas.aggregate(total=Sum('quilometragem'))['total'] or 0
    total_emissoes = viagens_filtradas.aggregate(total=Sum('Emissões_CO2'))['total'] or 0
    rpm_medio = viagens_filtradas.aggregate(media=Avg('RPM_médio'))['media'] or 0
    velocidade_media = viagens_filtradas.aggregate(media=Avg('Velocidade_média'))['media'] or 0
    media_quilometragem = viagens_filtradas.aggregate(media=Avg('Quilometragem_média'))['media'] or 0
    
    # Consumo com limite configurável
    consumo_max_normal = Config.consumo_maximo_normal()
    total_consumo = viagens_filtradas.filter(Consumido__lte=consumo_max_normal).aggregate(total=Sum('Consumido'))['total'] or 0
    
    # Calculadora de custos
    custo_diesel = Config.custo_diesel()
    
    # Calcular média real de eficiência dos dados filtrados
    media_km_atual_calculada = viagens_filtradas.filter(
        Consumido__gt=0, Consumido__lte=consumo_max_normal
    ).aggregate(media=Avg('Quilometragem_média'))['media']
    
    media_km_atual = float(media_km_atual_calculada) if media_km_atual_calculada and media_km_atual_calculada > 0 else Config.media_km_atual()
    media_km_objetivo = Config.media_km_objetivo()
    
    # Cálculos de economia
    km_atual_litros = float(total_quilometragem) / media_km_atual if media_km_atual > 0 else 0
    km_objetivo_litros = float(total_quilometragem) / media_km_objetivo if media_km_objetivo > 0 else 0
    
    custo_atual = km_atual_litros * custo_diesel
    custo_objetivo = km_objetivo_litros * custo_diesel
    economia_potencial = custo_atual - custo_objetivo
    
    # Estatísticas por marca (apenas da empresa logada)
    marcas_stats = {}
    marcas_disponiveis = viagens_filtradas.values_list('unidade__marca', flat=True).distinct()
    
    for marca in marcas_disponiveis:
        if marca:  # Evitar valores None
            stats = calcular_stats_marca_novo(viagens_filtradas, marca)
            stats['marca'] = marca
            marcas_stats[marca] = stats
    
    # Aplicar filtros na busca de unidades (apenas da empresa logada)
    unidades_filtradas = Unidade.objects.select_related('empresa').filter(empresa_id=empresa_logada_id)
    if classe_selecionada != 'todas':
        if classe_selecionada == 'veiculo':
            unidades_filtradas = unidades_filtradas.filter(cls__icontains='veículo')
        elif classe_selecionada == 'motorista':
            unidades_filtradas = unidades_filtradas.filter(cls__icontains='motorista')
    
    todas_unidades = []
    for unidade in unidades_filtradas.order_by('cls', 'id'):
        viagens_unidade_filtradas = viagens_filtradas.filter(unidade=unidade)
        stats = viagens_unidade_filtradas.aggregate(
            total_viagens=Count('id'),
            eficiencia_media=Avg('Quilometragem_média'),
            ultima_viagem_periodo=Max('período')
        )
        unidade.total_viagens = stats['total_viagens'] or 0
        unidade.eficiencia_media = stats['eficiencia_media']
        unidade.periodo_stats = stats['ultima_viagem_periodo']
        todas_unidades.append(unidade)
    
    # Context específico para empresa (sem dados de outras empresas)
    context = get_base_context_novo(empresa_selecionada, marca_selecionada, periodo_selecionado, filtro_combustivel, classe_selecionada, remover_zero)
    context.update({
        'empresa_logada': empresa_logada,
        'viagens_motoristas': viagens_motoristas,
        'viagens_veiculos': viagens_veiculos,
        'marcas_stats': marcas_stats,
        'todas_unidades': todas_unidades,
        'total_unidades': total_unidades_empresa,
        'total_quilometragem': total_quilometragem,
        'total_consumo': total_consumo,
        'total_emissoes': total_emissoes,
        'rpm_medio': rpm_medio,
        'velocidade_media': velocidade_media,
        'custo_diesel': custo_diesel,
        'media_km_objetivo': media_km_objetivo,
        'media_km_atual': media_km_atual,
        'custo_atual': custo_atual,
        'custo_objetivo': custo_objetivo,
        'economia_potencial': economia_potencial,
        'media_quilometragem': media_quilometragem,
        # Remover dados de outras empresas
        'empresas_stats': {},  # Vazio pois só tem uma empresa
        'total_empresas': 1,   # Sempre 1 para view da empresa
    })
    
    return render(request, 'umbrella360/report_empresa.html', context)

def ranking_empresa(request):
    """View para página de ranking da empresa logada"""
    # Verificar se há empresa logada
    empresa_logada_id = request.session.get('empresa_logada')
    if not empresa_logada_id:
        # Se não estiver logado, redirecionar para login
        return HttpResponseRedirect(reverse('login'))
    
    try:
        empresa_logada = Empresa.objects.get(id=empresa_logada_id)
    except Empresa.DoesNotExist:
        # Se empresa não existe mais, limpar sessão e redirecionar
        request.session.flush()
        return HttpResponseRedirect(reverse('login'))
    
    # Obter filtros opcionais
    periodo_selecionado = request.GET.get('periodo', 'todos')
    
    # Buscar viagens apenas da empresa logada
    viagens_base = Viagem_Base.objects.select_related('unidade', 'unidade__empresa').filter(
        unidade__empresa_id=empresa_logada_id
    )
    
    # Aplicar filtro de período se selecionado
    if periodo_selecionado != 'todos':
        viagens_base = viagens_base.filter(período=periodo_selecionado)
    
    # Filtrar viagens com eficiência realista (entre 1 e 4 km/L)
    viagens_filtradas = viagens_base.filter(
        Quilometragem_média__gte=1.0,
        Quilometragem_média__lte=4.0
    )
    
    # Top 10 Motoristas por eficiência
    top_motoristas = viagens_filtradas.filter(
        unidade__cls__icontains='motorista'
    ).order_by('-Quilometragem_média')
    
    # Top 10 Veículos por eficiência
    top_veiculos = viagens_filtradas.filter(
        unidade__cls__icontains='veículo'
    ).order_by('-Quilometragem_média')
    
    # Estatísticas gerais da empresa
    stats_empresa = viagens_filtradas.aggregate(
        total_unidades=Count('unidade', distinct=True),
        media_eficiencia=Avg('Quilometragem_média'),
        total_quilometragem=Sum('quilometragem'),
        total_consumo=Sum('Consumido'),
        total_viagens=Count('id')
    )
    
    # Contadores por tipo
    total_motoristas = viagens_filtradas.filter(unidade__cls__icontains='motorista').values('unidade').distinct().count()
    total_veiculos = viagens_filtradas.filter(unidade__cls__icontains='veículo').values('unidade').distinct().count()
    
    # Períodos disponíveis para filtro
    periodos_disponiveis = Viagem_Base.objects.filter(
        unidade__empresa_id=empresa_logada_id
    ).values_list('período', flat=True).distinct().order_by('período')
    
    # DADOS DOS CHECKPOINTS
    # Buscar checkpoints apenas da empresa logada
    checkpoints_base = CheckPoint.objects.select_related('unidade').filter(
        unidade__empresa_id=empresa_logada_id
    )
    
    # Aplicar filtro de período se selecionado
    if periodo_selecionado != 'todos':
        checkpoints_base = checkpoints_base.filter(período=periodo_selecionado)
    
    # Obter checkpoints ordenados por data de entrada mais recente
    checkpoints_recentes = checkpoints_base.order_by('-data_entrada')[:50]  # Últimos 50 registros
    
    # Estatísticas dos checkpoints
    total_checkpoints = checkpoints_base.count()
    cercas_ativas = checkpoints_base.values('cerca').distinct().count()
    unidades_com_checkpoint = checkpoints_base.values('unidade').distinct().count()
    
    # Cercas mais utilizadas
    cercas_populares = checkpoints_base.values('cerca').annotate(
        total_passagens=Count('id')
    ).order_by('-total_passagens')[:10]
    
    # Unidades mais ativas nos checkpoints
    unidades_ativas_checkpoint = checkpoints_base.values(
        'unidade__id', 'unidade__nm', 'unidade__cls'
    ).annotate(
        total_passagens=Count('id')
    ).order_by('-total_passagens')[:10]
    
    # DADOS DAS INFRAÇÕES
    # Buscar infrações apenas da empresa logada
    from .models import Infrações
    infracoes_base = Infrações.objects.select_related('unidade').filter(
        unidade__empresa_id=empresa_logada_id
    )
    
    # Obter infrações ordenadas por data mais recente
    infracoes_recentes = infracoes_base.order_by('-data')[:50]  # Últimas 50 infrações
    
    # Estatísticas das infrações
    total_infracoes = infracoes_base.count()
    unidades_com_infracoes = infracoes_base.values('unidade').distinct().count()
    
    # Velocidade média das infrações
    velocidade_media_infracoes = infracoes_base.aggregate(
        media_velocidade=Avg('velocidade'),
        media_limite=Avg('limite')
    )['media_velocidade'] or 0
    
    # Unidades com mais infrações
    unidades_mais_infracoes = infracoes_base.values(
        'unidade__id', 'unidade__nm', 'unidade__cls'
    ).annotate(
        total_infracoes=Count('id'),
        velocidade_maxima=Max('velocidade')
    ).order_by('-total_infracoes')[:10]
    
    # Infrações por tipo de limite (mais comuns)
    limites_mais_infringidos = infracoes_base.values('limite').annotate(
        total_ocorrencias=Count('id')
    ).order_by('-total_ocorrencias')[:10]
    
    context = {
        'empresa_logada': empresa_logada,
        'top_motoristas': top_motoristas,
        'top_veiculos': top_veiculos,
        'stats_empresa': stats_empresa,
        'total_motoristas': total_motoristas,
        'total_veiculos': total_veiculos,
        'periodo_selecionado': periodo_selecionado,
        'periodos_disponiveis': periodos_disponiveis,
        # Dados dos checkpoints
        'checkpoints_recentes': checkpoints_recentes,
        'total_checkpoints': total_checkpoints,
        'cercas_ativas': cercas_ativas,
        'unidades_com_checkpoint': unidades_com_checkpoint,
        'cercas_populares': cercas_populares,
        'unidades_ativas_checkpoint': unidades_ativas_checkpoint,
        # Dados das infrações
        'infracoes_recentes': infracoes_recentes,
        'total_infracoes': total_infracoes,
        'unidades_com_infracoes': unidades_com_infracoes,
        #'velocidade_media_infracoes': velocidade_media_infracoes['media_velocidade'] or 0,
        #'limite_medio_infracoes': velocidade_media_infracoes['media_limite'] or 0,
        'unidades_mais_infracoes': unidades_mais_infracoes,
        'limites_mais_infringidos': limites_mais_infringidos,
    }
    
    return render(request, 'umbrella360/ranking_empresa.html', context)


def lista_empresa(request):
    """View para listar todas as unidades de uma empresa logada"""
    empresas = Empresa.objects.filter(usuarios=request.user)
    return render(request, 'umbrella360/lista_empresa.html', {'empresas': empresas})

def logout_view(request):
    """View para logout da empresa"""
    request.session.flush()  # Remove todos os dados da sessão
    return HttpResponseRedirect(reverse('index'))

def generate_checkpoint_charts_data(checkpoints_queryset):
    """Gera dados para os gráficos de checkpoints"""
    from django.db.models import Sum, Count
    
    # 1. Tempo de permanência por cerca (versão simplificada)
    tempo_por_cerca = []
    try:
        # Buscar todas as cercas com checkpoints
        cercas = checkpoints_queryset.values_list('cerca', flat=True).distinct()[:10]
        
        for cerca in cercas:
            checkpoints_cerca = checkpoints_queryset.filter(cerca=cerca).exclude(duracao__isnull=True)
            
            if checkpoints_cerca.exists():
                tempo_total_segundos = 0
                count = 0
                
                for cp in checkpoints_cerca:
                    if cp.duracao:
                        tempo_total_segundos += cp.duracao.total_seconds()
                        count += 1
                
                if count > 0:
                    tempo_total_minutos = int(tempo_total_segundos / 60)
                    if tempo_total_minutos > 0:
                        tempo_por_cerca.append({
                            'cerca': cerca[:15] + '...' if len(cerca) > 15 else cerca,
                            'tempo_total_minutos': tempo_total_minutos
                        })
        
        # Ordenar por tempo total
        tempo_por_cerca = sorted(tempo_por_cerca, key=lambda x: x['tempo_total_minutos'], reverse=True)[:8]
    except Exception as e:
        print(f"Erro em tempo_por_cerca: {e}")
        tempo_por_cerca = []
    
    # 2. Número de passagens por cerca
    passagens_por_cerca = []
    try:
        cercas_count = {}
        for checkpoint in checkpoints_queryset:
            cerca = checkpoint.cerca
            if cerca:
                cerca_short = cerca[:15] + '...' if len(cerca) > 15 else cerca
                if cerca_short not in cercas_count:
                    cercas_count[cerca_short] = 0
                cercas_count[cerca_short] += 1
        
        # Converter para lista e pegar os top 8
        for cerca, count in sorted(cercas_count.items(), key=lambda x: x[1], reverse=True)[:8]:
            passagens_por_cerca.append({
                'cerca': cerca,
                'total_passagens': count
            })
    except Exception as e:
        print(f"Erro em passagens_por_cerca: {e}")
        passagens_por_cerca = []
    
    # 3. Atividade por hora do dia
    atividade_por_hora = []
    try:
        horas_count = {i: 0 for i in range(24)}
        
        for checkpoint in checkpoints_queryset:
            if checkpoint.data_entrada:
                hora = checkpoint.data_entrada.hour
                horas_count[hora] += 1
        
        atividade_por_hora = [
            {'hora': hora, 'total_atividade': total}
            for hora, total in horas_count.items()
        ]
    except Exception as e:
        print(f"Erro em atividade_por_hora: {e}")
        atividade_por_hora = []
    
    # 4. Atividade por dia da semana
    atividade_por_dia = []
    try:
        dias_count = {i: 0 for i in range(7)}  # 0=Segunda, 6=Domingo
        
        for checkpoint in checkpoints_queryset:
            if checkpoint.data_entrada:
                # Python: Monday=0, Sunday=6
                # Queremos: Monday=0, Sunday=6
                dia = checkpoint.data_entrada.weekday()
                dias_count[dia] += 1
        
        atividade_por_dia = [
            {'dia_semana': dia, 'total_atividade': total}
            for dia, total in dias_count.items()
        ]
    except Exception as e:
        print(f"Erro em atividade_por_dia: {e}")
        atividade_por_dia = []
    
    # 5. Timeline dos últimos 30 dias
    timeline = []
    try:
        from datetime import timedelta
        data_limite = datetime.now() - timedelta(days=30)
        
        datas_count = {}
        for checkpoint in checkpoints_queryset:
            if checkpoint.data_entrada and checkpoint.data_entrada >= data_limite:
                data_str = checkpoint.data_entrada.strftime('%Y-%m-%d')
                if data_str not in datas_count:
                    datas_count[data_str] = 0
                datas_count[data_str] += 1
        
        # Ordenar por data
        for data_str, count in sorted(datas_count.items()):
            timeline.append({
                'data': data_str,
                'total_checkpoints': count
            })
    except Exception as e:
        print(f"Erro em timeline: {e}")
        timeline = []
    
    # Debug: imprimir dados
    print(f"=== DADOS DOS GRÁFICOS ===")
    print(f"Tempo por cerca: {len(tempo_por_cerca)} items")
    print(f"Passagens por cerca: {len(passagens_por_cerca)} items") 
    print(f"Atividade por hora: {len(atividade_por_hora)} items")
    print(f"Atividade por dia: {len(atividade_por_dia)} items")
    print(f"Timeline: {len(timeline)} items")
    
    return {
        'tempoPerCerca': tempo_por_cerca,
        'passagensPerCerca': passagens_por_cerca,
        'atividadePerHora': atividade_por_hora,
        'atividadePerDia': atividade_por_dia,
        'timeline': timeline,
    }

def cercas_unidade(request, unidade_id):
    """View para exibir todos os checkpoints de uma unidade específica"""
    unidade = get_object_or_404(Unidade, id=unidade_id)
    
    # Verificar se usuário tem acesso à empresa da unidade
    empresa_logada_id = request.session.get('empresa_id')
    if empresa_logada_id and str(unidade.empresa_id) != str(empresa_logada_id):
        return HttpResponseRedirect(reverse('index'))
    
    # Query base de checkpoints
    checkpoints_queryset = CheckPoint.objects.filter(unidade=unidade).order_by('-data_entrada')
    
    # Aplicar filtros
    cerca_selecionada = request.GET.get('cerca', '')
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    tipo_evento = request.GET.get('tipo_evento', '')
    
    # Filtro por cerca
    if cerca_selecionada:
        checkpoints_queryset = checkpoints_queryset.filter(cerca__icontains=cerca_selecionada)
    
    # Filtro por data de início
    if data_inicio:
        try:
            data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d')
            checkpoints_queryset = checkpoints_queryset.filter(data_entrada__gte=data_inicio_dt)
        except ValueError:
            pass
    
    # Filtro por data de fim
    if data_fim:
        try:
            data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d')
            checkpoints_queryset = checkpoints_queryset.filter(data_entrada__lte=data_fim_dt)
        except ValueError:
            pass
    
    # Filtro por tipo de evento
    if tipo_evento == 'entrada':
        checkpoints_queryset = checkpoints_queryset.filter(data_entrada__isnull=False)
    elif tipo_evento == 'saida':
        checkpoints_queryset = checkpoints_queryset.filter(data_saida__isnull=False)
    elif tipo_evento == 'permanencia':
        checkpoints_queryset = checkpoints_queryset.filter(
            data_entrada__isnull=False, 
            data_saida__isnull=False,
            duracao__isnull=False
        )
    
    # Obter todos os checkpoints (sem limite)
    checkpoints = checkpoints_queryset
    
    # Estatísticas gerais
    total_checkpoints = checkpoints.count()
    cercas_distintas = checkpoints.values('cerca').distinct().count()
    
    # Calcular tempo total parado e média de permanência
    tempo_total_parado = "0 min"
    media_permanencia = "0 min"
    
    checkpoints_com_duracao = checkpoints.exclude(duracao__isnull=True)
    if checkpoints_com_duracao.exists():
        try:
            duracao_total_segundos = sum([
                cp.duracao.total_seconds() 
                for cp in checkpoints_com_duracao 
                if cp.duracao
            ])
            tempo_total_horas = int(duracao_total_segundos // 3600)
            tempo_total_minutos = int((duracao_total_segundos % 3600) // 60)
            
            if tempo_total_horas > 0:
                tempo_total_parado = f"{tempo_total_horas}h {tempo_total_minutos}min"
            else:
                tempo_total_parado = f"{tempo_total_minutos} min"
                
            # Média de permanência
            media_segundos = duracao_total_segundos / checkpoints_com_duracao.count()
            media_horas = int(media_segundos // 3600)
            media_minutos = int((media_segundos % 3600) // 60)
            
            if media_horas > 0:
                media_permanencia = f"{media_horas}h {media_minutos}min"
            else:
                media_permanencia = f"{media_minutos} min"
                
        except Exception:
            tempo_total_parado = "Erro no cálculo"
            media_permanencia = "Erro no cálculo"
    
    # Lista de cercas disponíveis para o filtro
    cercas_disponiveis = CheckPoint.objects.filter(unidade=unidade).values_list(
        'cerca', flat=True
    ).distinct().order_by('cerca')
    
    # Dados para gráficos
    chart_data = generate_checkpoint_charts_data(checkpoints_queryset)
    
    # Converter dados dos gráficos para JSON
    import json
    chart_data_json = json.dumps(chart_data, default=str)
    
    # Implementar paginação
    paginator = Paginator(checkpoints, 100)  # 100 registros por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'unidade': unidade,
        'checkpoints': page_obj,
        'total_checkpoints': total_checkpoints,
        'cercas_distintas': cercas_distintas,
        'tempo_total_parado': tempo_total_parado,
        'media_permanencia': media_permanencia,
        'cercas_disponiveis': cercas_disponiveis,
        'cerca_selecionada': cerca_selecionada,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'tipo_evento': tipo_evento,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        # Dados para gráficos em JSON
        'chart_data_json': chart_data_json,
    }
    
    return render(request, 'umbrella360/unidades/cercas.html', context)

def export_cercas_excel(request, unidade_id):
    """Exporta checkpoints de uma unidade para Excel"""
    unidade = get_object_or_404(Unidade, id=unidade_id)
    
    # Verificar se usuário tem acesso à empresa da unidade
    empresa_logada_id = request.session.get('empresa_id')
    if empresa_logada_id and str(unidade.empresa_id) != str(empresa_logada_id):
        return HttpResponseRedirect(reverse('index'))
    
    # Aplicar os mesmos filtros da view principal
    checkpoints_queryset = CheckPoint.objects.filter(unidade=unidade).order_by('-data_entrada')
    
    # Aplicar filtros
    cerca_selecionada = request.GET.get('cerca', '')
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    tipo_evento = request.GET.get('tipo_evento', '')
    
    if cerca_selecionada:
        checkpoints_queryset = checkpoints_queryset.filter(cerca__icontains=cerca_selecionada)
    
    if data_inicio:
        try:
            data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d')
            checkpoints_queryset = checkpoints_queryset.filter(data_entrada__gte=data_inicio_dt)
        except ValueError:
            pass
    
    if data_fim:
        try:
            data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d')
            checkpoints_queryset = checkpoints_queryset.filter(data_entrada__lte=data_fim_dt)
        except ValueError:
            pass
    
    if tipo_evento == 'entrada':
        checkpoints_queryset = checkpoints_queryset.filter(data_entrada__isnull=False)
    elif tipo_evento == 'saida':
        checkpoints_queryset = checkpoints_queryset.filter(data_saida__isnull=False)
    elif tipo_evento == 'permanencia':
        checkpoints_queryset = checkpoints_queryset.filter(
            data_entrada__isnull=False, 
            data_saida__isnull=False,
            duracao__isnull=False
        )
    
    # Criar workbook Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CheckPoints"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Cabeçalho do relatório
    ws.merge_cells('A1:H1')
    ws['A1'] = f"CheckPoints - {unidade.nm or unidade.id}"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Informações de filtro
    row = 2
    if cerca_selecionada:
        ws[f'A{row}'] = f"Filtro por Cerca: {cerca_selecionada}"
        row += 1
    if data_inicio:
        ws[f'A{row}'] = f"Data Início: {data_inicio}"
        row += 1
    if data_fim:
        ws[f'A{row}'] = f"Data Fim: {data_fim}"
        row += 1
    if tipo_evento:
        ws[f'A{row}'] = f"Tipo de Evento: {tipo_evento}"
        row += 1
    
    # Espaçamento
    row += 1
    
    # Cabeçalhos das colunas
    headers = ['Cerca', 'Data Entrada', 'Hora Entrada', 'Data Saída', 'Hora Saída', 'Duração', 'Status', 'Dia da Semana']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Dados
    for checkpoint in checkpoints_queryset:
        row += 1
        
        # Cerca
        ws.cell(row=row, column=1).value = checkpoint.cerca
        
        # Data e hora de entrada
        if checkpoint.data_entrada:
            ws.cell(row=row, column=2).value = checkpoint.data_entrada.strftime('%d/%m/%Y')
            ws.cell(row=row, column=3).value = checkpoint.data_entrada.strftime('%H:%M:%S')
        else:
            ws.cell(row=row, column=2).value = '-'
            ws.cell(row=row, column=3).value = '-'
        
        # Data e hora de saída
        if checkpoint.data_saida:
            ws.cell(row=row, column=4).value = checkpoint.data_saida.strftime('%d/%m/%Y')
            ws.cell(row=row, column=5).value = checkpoint.data_saida.strftime('%H:%M:%S')
        else:
            ws.cell(row=row, column=4).value = '-'
            ws.cell(row=row, column=5).value = '-'
        
        # Duração
        ws.cell(row=row, column=6).value = str(checkpoint.duracao) if checkpoint.duracao else '-'
        
        # Status
        if checkpoint.data_saida:
            status = 'Finalizada'
        elif checkpoint.data_entrada:
            status = 'Ativa'
        else:
            status = 'Iniciada'
        ws.cell(row=row, column=7).value = status
        
        # Dia da semana
        if checkpoint.data_entrada:
            dias_semana = {
                0: 'Segunda-feira', 1: 'Terça-feira', 2: 'Quarta-feira', 
                3: 'Quinta-feira', 4: 'Sexta-feira', 5: 'Sábado', 6: 'Domingo'
            }
            ws.cell(row=row, column=8).value = dias_semana.get(checkpoint.data_entrada.weekday(), '-')
        else:
            ws.cell(row=row, column=8).value = '-'
        
        # Aplicar bordas
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = border
    
    # Ajustar largura das colunas
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Preparar resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f"checkpoints_{unidade.nm or unidade.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Salvar workbook na resposta
    wb.save(response)
    
    return response

def viagem_diaria(request):
    """View para planilha de viagem diária - navegação entre dias disponíveis, agrupado por motorista com veículos discriminados"""
    from datetime import datetime, timedelta
    from django.db.models import Sum, Avg, F, Case, When, FloatField
    from collections import defaultdict
    
    # Verificar se há empresa logada
    empresa_logada_id = request.session.get('empresa_logada')
    empresa_logada = None
    if empresa_logada_id:
        try:
            empresa_logada = Empresa.objects.get(id=empresa_logada_id)
        except Empresa.DoesNotExist:
            pass
    
    # Obter parâmetro de data da URL (formato: YYYY-MM-DD)
    data_param = request.GET.get('data')
    
    # Buscar viagens detalhadas dos motoristas
    viagens_query = Viagem_Detalhada.objects.select_related('unidade', 'unidade__empresa')
    viagens_query = viagens_query.filter(unidade__cls__icontains='motorista')
    
    # Aplicar filtro de empresa se necessário
    if empresa_logada:
        viagens_query = viagens_query.filter(unidade__empresa=empresa_logada)
    
    # Obter todas as datas disponíveis nos dados (convertendo timestamps para datas)
    datas_disponiveis = set()
    for viagem in viagens_query.filter(timestamp_inicial__isnull=False):
        try:
            data_viagem = datetime.fromtimestamp(viagem.timestamp_inicial).date()
            datas_disponiveis.add(data_viagem)
        except (ValueError, OSError):
            continue
    
    datas_disponiveis = sorted(list(datas_disponiveis), reverse=True)  # Mais recente primeiro
    
    # Determinar qual data usar
    if data_param:
        try:
            data_selecionada = datetime.strptime(data_param, '%Y-%m-%d').date()
            # Verificar se a data está disponível
            if data_selecionada not in datas_disponiveis:
                data_selecionada = datas_disponiveis[0] if datas_disponiveis else timezone.now().date()
        except (ValueError, IndexError):
            data_selecionada = datas_disponiveis[0] if datas_disponiveis else timezone.now().date()
    else:
        # Usar a data mais recente disponível
        data_selecionada = datas_disponiveis[0] if datas_disponiveis else timezone.now().date()
    
    # Calcular timestamps do dia selecionado (início e fim do dia)
    inicio_dia = datetime.combine(data_selecionada, datetime.min.time())
    fim_dia = datetime.combine(data_selecionada, datetime.max.time())
    
    timestamp_inicio = int(inicio_dia.timestamp())
    timestamp_fim = int(fim_dia.timestamp())
    
    # Filtrar viagens do dia selecionado
    viagens_do_dia = viagens_query.filter(
        timestamp_inicial__gte=timestamp_inicio,
        timestamp_inicial__lte=timestamp_fim
    )
    
    # Aplicar filtros de qualidade dos dados
    viagens_do_dia = viagens_do_dia.filter(
        quilometragem__gt=0,
        Consumido__gt=0,
        Quilometragem_média__gte=1.0,
        Quilometragem_média__lte=5.0
    )
    
    # Agrupar por motorista E veículo
    viagens_agrupadas = viagens_do_dia.values(
        'unidade',
        'unidade__nm',
        'veiculo'
    ).annotate(
        total_quilometragem=Sum('quilometragem'),
        total_consumido=Sum('Consumido'),
        media_velocidade=Avg('Velocidade_média'),
        media_rpm=Avg('RPM_médio'),
        media_temperatura=Avg('Temperatura_média'),
        total_emissoes=Sum('Emissões_CO2'),
        eficiencia_media=Case(
            When(total_consumido__gt=0, then=F('total_quilometragem') / F('total_consumido')),
            default=0.0,
            output_field=FloatField()
        ),
        numero_viagens=Count('id')
    ).order_by('unidade__nm', 'veiculo')
    
    # 🔥 REORGANIZAR DADOS PARA AGRUPAMENTO VISUAL POR MOTORISTA
    motoristas_agrupados = defaultdict(list)
    
    for viagem in viagens_agrupadas:
        motorista_key = viagem['unidade__nm'] or viagem['unidade']
        motoristas_agrupados[motorista_key].append({
            'veiculo': viagem['veiculo'] or 'N/A',
            'total_quilometragem': viagem['total_quilometragem'],
            'total_consumido': viagem['total_consumido'],
            'eficiencia_media': viagem['eficiencia_media'],
            'numero_viagens': viagem['numero_viagens'],
            'media_velocidade': viagem['media_velocidade'],
        })
    
    # Transformar em lista ordenada para o template
    motoristas_lista = [
        {
            'nome': motorista,
            'veiculos': veiculos,
            'rowspan': len(veiculos)  # Número de linhas que o nome ocupará
        }
        for motorista, veiculos in sorted(motoristas_agrupados.items())
    ]
    
    # Calcular estatísticas do dia
    stats = {
        'total_unidades': len(motoristas_agrupados),  # Número único de motoristas
        'total_veiculos': viagens_agrupadas.values('veiculo').distinct().count(),
        'total_registros': viagens_agrupadas.count(),
        'total_km': sum(item['total_quilometragem'] for item in viagens_agrupadas if item['total_quilometragem']),
        'total_combustivel': sum(item['total_consumido'] for item in viagens_agrupadas if item['total_consumido']),
        'media_geral': (sum(item['total_quilometragem'] for item in viagens_agrupadas if item['total_quilometragem']) /
                        sum(item['total_consumido'] for item in viagens_agrupadas if item['total_consumido'])) if 
                       sum(item['total_consumido'] for item in viagens_agrupadas if item['total_consumido']) > 0 else 0
    }
    
    # Classificar por eficiência
    unidades_eficientes = sum(1 for motorista in motoristas_lista 
                              for veiculo in motorista['veiculos'] 
                              if veiculo['eficiencia_media'] >= 1.78)
    unidades_regulares = sum(1 for motorista in motoristas_lista 
                             for veiculo in motorista['veiculos'] 
                             if 1.5 <= veiculo['eficiencia_media'] < 1.78)
    unidades_ineficientes = sum(1 for motorista in motoristas_lista 
                                for veiculo in motorista['veiculos'] 
                                if veiculo['eficiencia_media'] < 1.5)
    
    # Navegação entre datas
    data_anterior = None
    data_proxima = None
    
    if data_selecionada in datas_disponiveis:
        indice_atual = datas_disponiveis.index(data_selecionada)
        
        if indice_atual > 0:
            data_anterior = datas_disponiveis[indice_atual - 1]
        
        if indice_atual < len(datas_disponiveis) - 1:
            data_proxima = datas_disponiveis[indice_atual + 1]
    
    context = {
        'motoristas': motoristas_lista,  # 🔥 Nova estrutura agrupada
        'data_selecionada': data_selecionada,
        'data_anterior': data_anterior,
        'data_proxima': data_proxima,
        'datas_disponiveis': datas_disponiveis[:10],
        'total_unidades': stats['total_unidades'],
        'total_veiculos': stats['total_veiculos'],
        'total_registros': stats['total_registros'],
        'total_km': stats['total_km'],
        'total_combustivel': stats['total_combustivel'],
        'media_geral': stats['media_geral'],
        'unidades_eficientes': unidades_eficientes,
        'unidades_regulares': unidades_regulares,
        'unidades_ineficientes': unidades_ineficientes,
        'empresa_logada': empresa_logada,
    }
    
    return render(request, 'umbrella360/Planilhas/viagem_diaria.html', context)

def export_viagem_diaria_excel(request):
    """
    Exporta relatório de viagem diária para Excel - agora com somatório por motorista E veículo
    """
    from datetime import datetime, timedelta
    from django.db.models import Sum, Avg, F, Case, When, FloatField, Count
    
    # ========== MESMA LÓGICA DA VIEW PRINCIPAL ==========
    # Verificar se há empresa logada
    empresa_logada_id = request.session.get('empresa_logada')
    empresa_logada = None
    if empresa_logada_id:
        try:
            empresa_logada = Empresa.objects.get(id=empresa_logada_id)
        except Empresa.DoesNotExist:
            pass

    # Obter parâmetro de data da URL
    data_param = request.GET.get('data')
    
    # ========== OBTENÇÃO DOS DADOS ==========
    viagens_query = Viagem_Detalhada.objects.select_related('unidade', 'unidade__empresa')
    viagens_query = viagens_query.filter(unidade__cls__icontains='motorista')
    
    if empresa_logada:
        viagens_query = viagens_query.filter(unidade__empresa=empresa_logada)
    
    # Obter todas as datas disponíveis
    datas_disponiveis = set()
    for viagem in viagens_query.filter(timestamp_inicial__isnull=False):
        try:
            data_viagem = datetime.fromtimestamp(viagem.timestamp_inicial).date()
            datas_disponiveis.add(data_viagem)
        except (ValueError, OSError):
            continue
    
    datas_disponiveis = sorted(list(datas_disponiveis), reverse=True)
    
    # Determinar qual data usar
    if data_param:
        try:
            data_selecionada = datetime.strptime(data_param, '%Y-%m-%d').date()
            if data_selecionada not in datas_disponiveis:
                data_selecionada = datas_disponiveis[0] if datas_disponiveis else timezone.now().date()
        except (ValueError, IndexError):
            data_selecionada = datas_disponiveis[0] if datas_disponiveis else timezone.now().date()
    else:
        data_selecionada = datas_disponiveis[0] if datas_disponiveis else timezone.now().date()
    
    # Calcular timestamps do dia selecionado
    inicio_dia = datetime.combine(data_selecionada, datetime.min.time())
    fim_dia = datetime.combine(data_selecionada, datetime.max.time())
    
    timestamp_inicio = int(inicio_dia.timestamp())
    timestamp_fim = int(fim_dia.timestamp())
    
    # Filtrar viagens do dia selecionado
    viagens_do_dia = viagens_query.filter(
        timestamp_inicial__gte=timestamp_inicio,
        timestamp_inicial__lte=timestamp_fim
    )
    
    # Aplicar filtros de qualidade
    viagens_do_dia = viagens_do_dia.filter(
        quilometragem__gt=0,
        Consumido__gt=0,
        Quilometragem_média__gte=1.0,
        Quilometragem_média__lte=5.0
    )
    
    # Agrupar por motorista E veículo
    viagens_agrupadas = viagens_do_dia.values(
        'unidade',
        'unidade__nm',
        'veiculo'  # INCLUÍDO NO AGRUPAMENTO
    ).annotate(
        total_quilometragem=Sum('quilometragem'),
        total_consumido=Sum('Consumido'),
        media_velocidade=Avg('Velocidade_média'),
        eficiencia_media=Case(
            When(total_consumido__gt=0, then=F('total_quilometragem') / F('total_consumido')),
            default=0.0,
            output_field=FloatField()
        ),
        numero_viagens=Count('id')
    ).order_by('unidade__nm', 'veiculo')
    
    # Calcular estatísticas
    stats = {
        'total_unidades': viagens_agrupadas.values('unidade').distinct().count(),
        'total_veiculos': viagens_agrupadas.values('veiculo').distinct().count(),
        'total_registros': viagens_agrupadas.count(),
        'total_km': sum(item['total_quilometragem'] for item in viagens_agrupadas if item['total_quilometragem']),
        'total_combustivel': sum(item['total_consumido'] for item in viagens_agrupadas if item['total_consumido']),
        'media_geral': (sum(item['total_quilometragem'] for item in viagens_agrupadas if item['total_quilometragem']) /
                        sum(item['total_consumido'] for item in viagens_agrupadas if item['total_consumido'])) if 
                       sum(item['total_consumido'] for item in viagens_agrupadas if item['total_consumido']) > 0 else 0
    }
    
    # ========== CRIAÇÃO DO EXCEL ==========
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Viagem Diária"
    
    # ========== ESTILOS ==========
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    title_font = Font(bold=True, size=16, color="2c3e50")
    stats_font = Font(bold=True, color="2c3e50")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========== CABEÇALHO PRINCIPAL ==========
    ws.merge_cells('A1:G1')
    ws['A1'] = f"📋 Relatório de Viagem Diária - {data_selecionada.strftime('%d/%m/%Y')} (Por Motorista e Veículo)"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Informações gerais
    ws.merge_cells('A2:G2')
    ws['A2'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    if empresa_logada:
        ws.merge_cells('A3:G3')
        ws['A3'] = f"Empresa: {empresa_logada.nome}"
        ws['A3'].alignment = Alignment(horizontal='center')
        start_row = 5
    else:
        start_row = 4
    
    # ========== ESTATÍSTICAS RESUMO ==========
    ws.merge_cells(f'A{start_row}:G{start_row}')
    ws[f'A{start_row}'] = "📊 RESUMO EXECUTIVO"
    ws[f'A{start_row}'].font = stats_font
    ws[f'A{start_row}'].alignment = Alignment(horizontal='center')
    
    stats_row = start_row + 1
    
    # Linha 1 de estatísticas
    ws[f'A{stats_row}'] = "Motoristas:"
    ws[f'B{stats_row}'] = stats['total_unidades']
    ws[f'C{stats_row}'] = "Veículos:"
    ws[f'D{stats_row}'] = stats['total_veiculos']
    ws[f'E{stats_row}'] = "Registros:"
    ws[f'F{stats_row}'] = stats['total_registros']
    
    # Linha 2 de estatísticas
    stats_row += 1
    ws[f'A{stats_row}'] = "KM Total:"
    ws[f'B{stats_row}'] = round(stats['total_km'], 2)
    ws[f'C{stats_row}'] = "Combustível:"
    ws[f'D{stats_row}'] = f"{round(stats['total_combustivel'], 2)} L"
    ws[f'E{stats_row}'] = "Média Geral:"
    ws[f'F{stats_row}'] = f"{round(stats['media_geral'], 2)} km/L"
    
    # ========== CABEÇALHOS DA TABELA ==========
    table_start = stats_row + 3
    
    headers = ['Motorista', 'Veículo', 'KM Total', 'Consumo Total (L)', 'Eficiência (km/L)', 'Viagens', 'Observações']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=table_start, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # ========== DADOS DA TABELA ==========
    current_row = table_start + 1
    
    for viagem in viagens_agrupadas:
        # Motorista
        ws.cell(row=current_row, column=1).value = viagem['unidade__nm'] or viagem['unidade']
        
        # Veículo
        ws.cell(row=current_row, column=2).value = viagem['veiculo'] or 'N/A'
        
        # KM Total
        ws.cell(row=current_row, column=3).value = round(viagem['total_quilometragem'], 2)
        
        # Consumo Total
        ws.cell(row=current_row, column=4).value = round(viagem['total_consumido'], 1)
        
        # Eficiência
        ws.cell(row=current_row, column=5).value = round(viagem['eficiencia_media'], 2)
        
        # Número de Viagens
        ws.cell(row=current_row, column=6).value = viagem['numero_viagens']
        
        # Observações
        eficiencia = viagem['eficiencia_media']
        if eficiencia >= 2.0:
            obs = "Excelente"
        elif eficiencia >= 1.78:
            obs = "Boa"
        elif eficiencia >= 1.5:
            obs = "Regular"
        else:
            obs = "Ruim - Requer atenção"
            
        ws.cell(row=current_row, column=7).value = obs
        
        # Aplicar bordas
        for col in range(1, 8):
            ws.cell(row=current_row, column=col).border = border
            
        current_row += 1
    
    # ========== CONFIGURAÇÕES FINAIS ==========
    column_widths = [25, 20, 12, 15, 15, 10, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # ========== RESPOSTA HTTP ==========
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"viagem_diaria_detalhada_{data_selecionada.strftime('%Y%m%d')}_{timestamp}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    
    return response
# ...existing code...

def export_cercas_pdf(request, unidade_id):
    """Exporta checkpoints de uma unidade para PDF"""
    unidade = get_object_or_404(Unidade, id=unidade_id)
    
    # Verificar se usuário tem acesso à empresa da unidade
    empresa_logada_id = request.session.get('empresa_id')
    if empresa_logada_id and str(unidade.empresa_id) != str(empresa_logada_id):
        return HttpResponseRedirect(reverse('index'))
    
    # Aplicar os mesmos filtros da view principal
    checkpoints_queryset = CheckPoint.objects.filter(unidade=unidade).order_by('-data_entrada')
    
    # Aplicar filtros
    cerca_selecionada = request.GET.get('cerca', '')
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    tipo_evento = request.GET.get('tipo_evento', '')
    
    if cerca_selecionada:
        checkpoints_queryset = checkpoints_queryset.filter(cerca__icontains=cerca_selecionada)
    
    if data_inicio:
        try:
            data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d')
            checkpoints_queryset = checkpoints_queryset.filter(data_entrada__gte=data_inicio_dt)
        except ValueError:
            pass
    
    if data_fim:
        try:
            data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d')
            checkpoints_queryset = checkpoints_queryset.filter(data_entrada__lte=data_fim_dt)
        except ValueError:
            pass
    
    if tipo_evento == 'entrada':
        checkpoints_queryset = checkpoints_queryset.filter(data_entrada__isnull=False)
    elif tipo_evento == 'saida':
        checkpoints_queryset = checkpoints_queryset.filter(data_saida__isnull=False)
    elif tipo_evento == 'permanencia':
        checkpoints_queryset = checkpoints_queryset.filter(
            data_entrada__isnull=False, 
            data_saida__isnull=False,
            duracao__isnull=False
        )
    
    # Criar buffer para o PDF
    buffer = io.BytesIO()
    
    # Criar documento PDF
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=18,
        spaceAfter=30,
        textColor=colors.darkblue,
        alignment=1  # Center alignment
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=12,
        spaceAfter=10,
        textColor=colors.grey,
        alignment=1
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=6
    )
    
    # Elementos do documento
    elements = []
    
    # Título
    title = Paragraph(f"CheckPoints - {unidade.nm or unidade.id}", title_style)
    elements.append(title)
    
    # Subtitle com data de geração
    subtitle = Paragraph(f"Relatório gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}", subtitle_style)
    elements.append(subtitle)
    
    # Informações de filtro
    filter_info = []
    if cerca_selecionada:
        filter_info.append(f"<b>Filtro por Cerca:</b> {cerca_selecionada}")
    if data_inicio:
        filter_info.append(f"<b>Data Início:</b> {data_inicio}")
    if data_fim:
        filter_info.append(f"<b>Data Fim:</b> {data_fim}")
    if tipo_evento:
        filter_info.append(f"<b>Tipo de Evento:</b> {tipo_evento}")
    
    if filter_info:
        elements.append(Spacer(1, 12))
        elements.append(Paragraph("<b>Filtros Aplicados:</b>", normal_style))
        for info in filter_info:
            elements.append(Paragraph(info, normal_style))
        elements.append(Spacer(1, 12))
    
    # Estatísticas
    total_checkpoints = checkpoints_queryset.count()
    elements.append(Paragraph(f"<b>Total de CheckPoints:</b> {total_checkpoints}", normal_style))
    elements.append(Spacer(1, 20))
    
    # Tabela de dados
    if checkpoints_queryset.exists():
        # Cabeçalhos da tabela
        table_data = [
            ['Cerca', 'Data Entrada', 'Data Saída', 'Duração', 'Status']
        ]
        
        # Dados
        for checkpoint in checkpoints_queryset:
            cerca = checkpoint.cerca[:25] + '...' if len(checkpoint.cerca) > 25 else checkpoint.cerca
            
            data_entrada = checkpoint.data_entrada.strftime('%d/%m/%Y %H:%M') if checkpoint.data_entrada else '-'
            data_saida = checkpoint.data_saida.strftime('%d/%m/%Y %H:%M') if checkpoint.data_saida else '-'
            duracao = str(checkpoint.duracao) if checkpoint.duracao else '-'
            
            if checkpoint.data_saida:
                status = 'Finalizada'
            elif checkpoint.data_entrada:
                status = 'Ativa'
            else:
                status = 'Iniciada'
            
            table_data.append([cerca, data_entrada, data_saida, duracao, status])
        
        # Criar tabela
        table = Table(table_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1*inch, 1*inch])
        
        # Estilo da tabela
        table.setStyle(TableStyle([
            # Cabeçalho
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            
            # Corpo da tabela
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            
            # Bordas
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Alternância de cores nas linhas
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        elements.append(table)
    else:
        elements.append(Paragraph("Nenhum checkpoint encontrado para os filtros aplicados.", normal_style))
    
    # Rodapé
    elements.append(Spacer(1, 30))
    footer = Paragraph("Relatório gerado pelo Sistema Umbrella360", subtitle_style)
    elements.append(footer)
    
    # Gerar PDF
    doc.build(elements)
    
    # Preparar resposta HTTP
    buffer.seek(0)
    filename = f"checkpoints_{unidade.nm or unidade.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response




def performance_frota(request):
    """View para análise de performance de RPM da frota toda - USANDO CAMPOS BOOLEANOS DO MODEL"""
    from datetime import datetime, timedelta
    from django.db.models import Count, Avg, Q
    
    # Verificar se há empresa logada
    empresa_logada_id = request.session.get('empresa_logada')
    empresa_logada = None
    
    if empresa_logada_id:
        try:
            empresa_logada = Empresa.objects.get(id=empresa_logada_id)
        except Empresa.DoesNotExist:
            return HttpResponseRedirect(reverse('login_view'))
    
    # 🔥 Obter filtros de data
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    periodo_dias = request.GET.get('periodo_dias', '30')  # Padrão: últimos 30 dias
    
    # Buscar viagens eco da empresa
    from .models import Viagem_eco
    
    viagens_eco_query = Viagem_eco.objects.select_related('unidade', 'unidade__empresa')
    
    if empresa_logada_id:
        viagens_eco_query = viagens_eco_query.filter(unidade__empresa_id=empresa_logada_id)
    
    # Aplicar filtros de data
    if data_inicio and data_fim:
        try:
            dt_inicio = datetime.strptime(data_inicio, '%Y-%m-%d')
            dt_fim = datetime.strptime(data_fim, '%Y-%m-%d')
            timestamp_inicio = int(dt_inicio.timestamp())
            timestamp_fim = int(dt_fim.replace(hour=23, minute=59, second=59).timestamp())
            viagens_eco_query = viagens_eco_query.filter(
                timestamp__gte=timestamp_inicio,
                timestamp__lte=timestamp_fim
            )
        except ValueError:
            pass
    elif periodo_dias and periodo_dias != 'todos':
        try:
            dias = int(periodo_dias)
            data_limite = datetime.now() - timedelta(days=dias)
            timestamp_limite = int(data_limite.timestamp())
            viagens_eco_query = viagens_eco_query.filter(timestamp__gte=timestamp_limite)
        except ValueError:
            pass

    # 🔥 NOVO: Processar dados agregados por unidade usando os campos booleanos
    unidades_performance = []
    
    if viagens_eco_query.exists():
        # Agrupar por unidade e contar classificações usando os campos booleanos
        unidades_stats = viagens_eco_query.values('unidade_id').annotate(
            total_pontos=Count('id'),
            
            # Contar cada faixa usando os campos booleanos
            faixa_azul_count=Count('id', filter=Q(faixa_azul=True)),
            faixa_verde_count=Count('id', filter=Q(faixa_verde=True)),
            faixa_amarela_count=Count('id', filter=Q(faixa_amarela=True)),
            faixa_vermelha_count=Count('id', filter=Q(faixa_vermelha=True)),
            motor_ocioso_count=Count('id', filter=Q(ocioso=True)),
            
            # Médias de RPM e velocidade (filtrar rpm > 0 para não distorcer)
            rpm_medio=Avg('rpm', filter=Q(rpm__gt=0)),
            rpm_maximo=Max('rpm', filter=Q(rpm__gt=0)),
            rpm_minimo=Min('rpm', filter=Q(rpm__gt=0)),
        )
        
        # Processar cada unidade
        for stats in unidades_stats:
            try:
                unidade = Unidade.objects.select_related('empresa').get(id=stats['unidade_id'])
                
                total_pontos = stats['total_pontos']
                
                # 🔥 CORREÇÃO: Calcular soma das classificações válidas (que somam 100%)
                # Total de pontos classificados = azul + verde + amarela + vermelha + ocioso
                total_classificados = (
                    stats['faixa_azul_count'] + 
                    stats['faixa_verde_count'] + 
                    stats['faixa_amarela_count'] + 
                    stats['faixa_vermelha_count'] + 
                    stats['motor_ocioso_count']
                )
                
                # Usar total_classificados como base para percentuais (DEVE SOMAR 100%)
                base_percentual = total_classificados if total_classificados > 0 else 1
                
                perc_azul = (stats['faixa_azul_count'] / base_percentual * 100)
                perc_verde = (stats['faixa_verde_count'] / base_percentual * 100)
                perc_amarela = (stats['faixa_amarela_count'] / base_percentual * 100)
                perc_vermelha = (stats['faixa_vermelha_count'] / base_percentual * 100)
                perc_ocioso = (stats['motor_ocioso_count'] / base_percentual * 100)
                
                # 🔥 VALIDAÇÃO: Garantir que soma seja ~100% (aceitar pequenas variações de arredondamento)
                soma_percentuais = perc_azul + perc_verde + perc_amarela + perc_vermelha + perc_ocioso
                # Se houver diferença por arredondamento, ajustar a maior categoria
                if abs(soma_percentuais - 100.0) > 0.1:
                    print(f"⚠️ AVISO: Soma dos percentuais para {unidade.nm} = {soma_percentuais:.2f}% (esperado: 100%)")
                
                # Score de performance (mais verde e azul = melhor, motor ocioso penaliza)
                score = (perc_verde * 1.0) + (perc_azul * 0.8) - (perc_amarela * 0.5) - (perc_vermelha * 1.0) - (perc_ocioso * 0.3)
                
                # Ajustar RPM (dividir por 8 conforme sua lógica)
                rpm_medio = (stats['rpm_medio'] / 8) if stats['rpm_medio'] else 0
                rpm_maximo = (stats['rpm_maximo'] / 8) if stats['rpm_maximo'] else 0
                rpm_minimo = (stats['rpm_minimo'] / 8) if stats['rpm_minimo'] else 0
                
                unidades_performance.append({
                    'unidade': unidade,
                    'total_pontos': total_pontos,
                    'total_classificados': total_classificados,  # 🔥 NOVO: Mostrar quantos foram classificados
                    'total_pontos_rpm': total_pontos - stats['motor_ocioso_count'],  # Pontos com RPM válido
                    'rpm_medio': rpm_medio,
                    'rpm_maximo': rpm_maximo,
                    'rpm_minimo': rpm_minimo,
                    'faixa_azul': stats['faixa_azul_count'],
                    'faixa_verde': stats['faixa_verde_count'],
                    'faixa_amarela': stats['faixa_amarela_count'],
                    'faixa_vermelha': stats['faixa_vermelha_count'],
                    'motor_ocioso': stats['motor_ocioso_count'],
                    'perc_azul': perc_azul,
                    'perc_verde': perc_verde,
                    'perc_amarela': perc_amarela,
                    'perc_vermelha': perc_vermelha,
                    'perc_ocioso': perc_ocioso,
                    'soma_percentuais': soma_percentuais,  # 🔥 NOVO: Para validação
                    'score': score,
                })
            except Unidade.DoesNotExist:
                continue
    
    # Ordenar por score (melhor performance primeiro)
    unidades_performance = sorted(unidades_performance, key=lambda x: x['score'], reverse=True)
    
    # Calcular estatísticas gerais da frota
    if unidades_performance:
        # 🔥 CORREÇÃO: Usar total_classificados para médias ponderadas corretas
        total_classificados_frota = sum(u['total_classificados'] for u in unidades_performance)
        
        stats_gerais = {
            'total_unidades': len(unidades_performance),
            'total_pontos': sum(u['total_pontos'] for u in unidades_performance),
            'total_classificados': total_classificados_frota,
            'rpm_medio_frota': sum(u['rpm_medio'] * u['total_pontos'] for u in unidades_performance) / sum(u['total_pontos'] for u in unidades_performance) if sum(u['total_pontos'] for u in unidades_performance) > 0 else 0,
            
            # 🔥 CORREÇÃO: Percentuais da frota somam 100%
            'perc_verde_frota': (sum(u['faixa_verde'] for u in unidades_performance) / total_classificados_frota * 100) if total_classificados_frota > 0 else 0,
            'perc_vermelha_frota': (sum(u['faixa_vermelha'] for u in unidades_performance) / total_classificados_frota * 100) if total_classificados_frota > 0 else 0,
            'perc_azul_frota': (sum(u['faixa_azul'] for u in unidades_performance) / total_classificados_frota * 100) if total_classificados_frota > 0 else 0,
            'perc_amarela_frota': (sum(u['faixa_amarela'] for u in unidades_performance) / total_classificados_frota * 100) if total_classificados_frota > 0 else 0,
            'perc_ocioso_frota': (sum(u['motor_ocioso'] for u in unidades_performance) / total_classificados_frota * 100) if total_classificados_frota > 0 else 0,
        }
        
        # 🔥 VALIDAÇÃO: Verificar se soma é ~100%
        soma_frota = (
            stats_gerais['perc_verde_frota'] + 
            stats_gerais['perc_vermelha_frota'] + 
            stats_gerais['perc_azul_frota'] + 
            stats_gerais['perc_amarela_frota'] + 
            stats_gerais['perc_ocioso_frota']
        )
        if abs(soma_frota - 100.0) > 0.1:
            print(f"⚠️ AVISO: Soma dos percentuais da frota = {soma_frota:.2f}% (esperado: 100%)")
    else:
        stats_gerais = {
            'total_unidades': 0,
            'total_pontos': 0,
            'total_classificados': 0,
            'rpm_medio_frota': 0,
            'perc_verde_frota': 0,
            'perc_vermelha_frota': 0,
            'perc_azul_frota': 0,
            'perc_amarela_frota': 0,
            'perc_ocioso_frota': 0,
        }
    
    context = {
        'empresa_logada': empresa_logada,
        'unidades_performance': unidades_performance,
        'stats_gerais': stats_gerais,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'periodo_dias': periodo_dias,
    }
    
    return render(request, 'umbrella360/performance.html', context)

def performance_motoristas(request):
    """View para análise de performance de RPM dos motoristas - USANDO CAMPOS BOOLEANOS DO MODEL"""
    from datetime import datetime, timedelta
    from django.db.models import Count, Avg, Q
    
    # Verificar se há empresa logada
    empresa_logada_id = request.session.get('empresa_logada')
    empresa_logada = None
    
    if empresa_logada_id:
        try:
            empresa_logada = Empresa.objects.get(id=empresa_logada_id)
        except Empresa.DoesNotExist:
            return HttpResponseRedirect(reverse('login_view'))
    
    # 🔥 Obter filtros de data
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    periodo_dias = request.GET.get('periodo_dias', '30')  # Padrão: últimos 30 dias
    
    # Buscar viagens eco relacionadas a motoristas
    from .models import Viagem_eco, Driver
    
    viagens_eco_query = Viagem_eco.objects.select_related('nome_motorista', 'unidade', 'unidade__empresa')
    
    # Filtrar apenas viagens que têm motorista associado
    viagens_eco_query = viagens_eco_query.filter(nome_motorista__isnull=False)
    
    if empresa_logada_id:
        viagens_eco_query = viagens_eco_query.filter(unidade__empresa_id=empresa_logada_id)
    
    # Aplicar filtros de data
    if data_inicio and data_fim:
        try:
            dt_inicio = datetime.strptime(data_inicio, '%Y-%m-%d')
            dt_fim = datetime.strptime(data_fim, '%Y-%m-%d')
            timestamp_inicio = int(dt_inicio.timestamp())
            timestamp_fim = int(dt_fim.replace(hour=23, minute=59, second=59).timestamp())
            viagens_eco_query = viagens_eco_query.filter(
                timestamp__gte=timestamp_inicio,
                timestamp__lte=timestamp_fim
            )
        except ValueError:
            pass
    elif periodo_dias and periodo_dias != 'todos':
        try:
            dias = int(periodo_dias)
            data_limite = datetime.now() - timedelta(days=dias)
            timestamp_limite = int(data_limite.timestamp())
            viagens_eco_query = viagens_eco_query.filter(timestamp__gte=timestamp_limite)
        except ValueError:
            pass

    # 🔥 NOVO: Processar dados por motorista usando campos booleanos
    motoristas_performance = []
    
    if viagens_eco_query.exists():
        # Agrupar por motorista e contar classificações usando os campos booleanos
        motoristas_stats = viagens_eco_query.values('nome_motorista_id').annotate(
            total_pontos=Count('id'),
            
            # Contar cada faixa usando os campos booleanos
            faixa_azul_count=Count('id', filter=Q(faixa_azul=True)),
            faixa_verde_count=Count('id', filter=Q(faixa_verde=True)),
            faixa_amarela_count=Count('id', filter=Q(faixa_amarela=True)),
            faixa_vermelha_count=Count('id', filter=Q(faixa_vermelha=True)),
            motor_ocioso_count=Count('id', filter=Q(ocioso=True)),
            
            # Médias de RPM e velocidade
            rpm_medio=Avg('rpm', filter=Q(rpm__gt=0)),
            rpm_maximo=Max('rpm', filter=Q(rpm__gt=0)),
            rpm_minimo=Min('rpm', filter=Q(rpm__gt=0)),
            velocidade_media=Avg('velocidade', filter=Q(velocidade__gt=0)),
            
            # Contar veículos únicos
            total_veiculos=Count('unidade_id', distinct=True)
        )
        
        # Processar cada motorista
        for stats in motoristas_stats:
            try:
                motorista = Driver.objects.select_related('empresa').get(id=stats['nome_motorista_id'])
                
                total_pontos = stats['total_pontos']
                
                # 🔥 CORREÇÃO: Calcular soma das classificações válidas (que somam 100%)
                total_classificados = (
                    stats['faixa_azul_count'] + 
                    stats['faixa_verde_count'] + 
                    stats['faixa_amarela_count'] + 
                    stats['faixa_vermelha_count'] + 
                    stats['motor_ocioso_count']
                )
                
                # Usar total_classificados como base para percentuais (DEVE SOMAR 100%)
                base_percentual = total_classificados if total_classificados > 0 else 1
                
                perc_azul = (stats['faixa_azul_count'] / base_percentual * 100)
                perc_verde = (stats['faixa_verde_count'] / base_percentual * 100)
                perc_amarela = (stats['faixa_amarela_count'] / base_percentual * 100)
                perc_vermelha = (stats['faixa_vermelha_count'] / base_percentual * 100)
                perc_ocioso = (stats['motor_ocioso_count'] / base_percentual * 100)
                
                # 🔥 VALIDAÇÃO: Garantir que soma seja ~100%
                soma_percentuais = perc_azul + perc_verde + perc_amarela + perc_vermelha + perc_ocioso
                if abs(soma_percentuais - 100.0) > 0.1:
                    print(f"⚠️ AVISO: Soma dos percentuais para {motorista.nm} = {soma_percentuais:.2f}% (esperado: 100%)")
                
                # Score de performance (mais verde e azul = melhor, motor ocioso penaliza)
                score = (perc_verde * 1.0) + (perc_azul * 0.8) - (perc_amarela * 0.5) - (perc_vermelha * 1.0) - (perc_ocioso * 0.3)
                
                # Ajustar RPM (dividir por 8 conforme sua lógica)
                rpm_medio = (stats['rpm_medio'] / 8) if stats['rpm_medio'] else 0
                rpm_maximo = (stats['rpm_maximo'] / 8) if stats['rpm_maximo'] else 0
                rpm_minimo = (stats['rpm_minimo'] / 8) if stats['rpm_minimo'] else 0
                
                motoristas_performance.append({
                    'motorista': motorista,
                    'total_pontos': total_pontos,
                    'total_classificados': total_classificados,
                    'total_veiculos': stats['total_veiculos'],
                    'total_pontos_rpm': total_pontos - stats['motor_ocioso_count'],
                    'rpm_medio': rpm_medio,
                    'rpm_maximo': rpm_maximo,
                    'rpm_minimo': rpm_minimo,
                    'velocidade_media': stats['velocidade_media'] or 0,
                    'faixa_azul': stats['faixa_azul_count'],
                    'faixa_verde': stats['faixa_verde_count'],
                    'faixa_amarela': stats['faixa_amarela_count'],
                    'faixa_vermelha': stats['faixa_vermelha_count'],
                    'motor_ocioso': stats['motor_ocioso_count'],
                    'perc_azul': perc_azul,
                    'perc_verde': perc_verde,
                    'perc_amarela': perc_amarela,
                    'perc_vermelha': perc_vermelha,
                    'perc_ocioso': perc_ocioso,
                    'soma_percentuais': soma_percentuais,
                    'score': score,
                })
            except Driver.DoesNotExist:
                continue
    
    # Ordenar por score (melhor performance primeiro)
    motoristas_performance = sorted(motoristas_performance, key=lambda x: x['score'], reverse=True)
    
    # Calcular estatísticas gerais
    if motoristas_performance:
        # 🔥 CORREÇÃO: Usar total_classificados para médias ponderadas corretas
        total_classificados_geral = sum(m['total_classificados'] for m in motoristas_performance)
        total_pontos_geral = sum(m['total_pontos'] for m in motoristas_performance)
        
        stats_gerais = {
            'total_motoristas': len(motoristas_performance),
            'total_pontos': total_pontos_geral,
            'total_classificados': total_classificados_geral,
            'rpm_medio_geral': sum(m['rpm_medio'] * m['total_pontos'] for m in motoristas_performance) / total_pontos_geral if total_pontos_geral > 0 else 0,
            'velocidade_media_geral': sum(m['velocidade_media'] * m['total_pontos'] for m in motoristas_performance) / total_pontos_geral if total_pontos_geral > 0 else 0,
            
            # 🔥 CORREÇÃO: Percentuais gerais somam 100%
            'perc_verde_geral': (sum(m['faixa_verde'] for m in motoristas_performance) / total_classificados_geral * 100) if total_classificados_geral > 0 else 0,
            'perc_vermelha_geral': (sum(m['faixa_vermelha'] for m in motoristas_performance) / total_classificados_geral * 100) if total_classificados_geral > 0 else 0,
            'perc_azul_geral': (sum(m['faixa_azul'] for m in motoristas_performance) / total_classificados_geral * 100) if total_classificados_geral > 0 else 0,
            'perc_amarela_geral': (sum(m['faixa_amarela'] for m in motoristas_performance) / total_classificados_geral * 100) if total_classificados_geral > 0 else 0,
            'perc_ocioso_geral': (sum(m['motor_ocioso'] for m in motoristas_performance) / total_classificados_geral * 100) if total_classificados_geral > 0 else 0,
        }
        
        # 🔥 VALIDAÇÃO: Verificar se soma é ~100%
        soma_geral = (
            stats_gerais['perc_verde_geral'] + 
            stats_gerais['perc_vermelha_geral'] + 
            stats_gerais['perc_azul_geral'] + 
            stats_gerais['perc_amarela_geral'] + 
            stats_gerais['perc_ocioso_geral']
        )
        if abs(soma_geral - 100.0) > 0.1:
            print(f"⚠️ AVISO: Soma dos percentuais gerais = {soma_geral:.2f}% (esperado: 100%)")
    else:
        stats_gerais = {
            'total_motoristas': 0,
            'total_pontos': 0,
            'total_classificados': 0,
            'rpm_medio_geral': 0,
            'velocidade_media_geral': 0,
            'perc_verde_geral': 0,
            'perc_vermelha_geral': 0,
            'perc_azul_geral': 0,
            'perc_amarela_geral': 0,
            'perc_ocioso_geral': 0,
        }
    
    # 🔥 NOVO: Preparar dados temporais usando campos booleanos
    from collections import defaultdict
    from datetime import date
    
    evolucao_temporal = []
    
    if viagens_eco_query.exists():
        # Agrupar viagens por data
        dados_por_data = defaultdict(lambda: {
            'total_registros': 0,
            'faixa_azul': 0,
            'faixa_verde': 0,
            'faixa_amarela': 0,
            'faixa_vermelha': 0,
            'motor_ocioso': 0,
            'rpms': [],
            'velocidades': []
        })
        
        for v in viagens_eco_query:
            try:
                data_viagem = date.fromtimestamp(v.timestamp)
                
                dados_por_data[data_viagem]['total_registros'] += 1
                
                # Contar classificações usando os campos booleanos
                if v.faixa_azul:
                    dados_por_data[data_viagem]['faixa_azul'] += 1
                if v.faixa_verde:
                    dados_por_data[data_viagem]['faixa_verde'] += 1
                if v.faixa_amarela:
                    dados_por_data[data_viagem]['faixa_amarela'] += 1
                if v.faixa_vermelha:
                    dados_por_data[data_viagem]['faixa_vermelha'] += 1
                if v.ocioso:
                    dados_por_data[data_viagem]['motor_ocioso'] += 1
                
                # Coletar RPM e velocidade para médias
                if v.rpm and v.rpm > 0:
                    dados_por_data[data_viagem]['rpms'].append(v.rpm / 8)
                if v.velocidade and v.velocidade > 0:
                    dados_por_data[data_viagem]['velocidades'].append(v.velocidade)
                    
            except (ValueError, OSError):
                continue
        
        # Calcular métricas por data
        for data, dados in sorted(dados_por_data.items()):
            total = dados['total_registros']
            if total > 0:
                # 🔥 Usar total como base (soma = 100%)
                perc_azul = (dados['faixa_azul'] / total * 100)
                perc_verde = (dados['faixa_verde'] / total * 100)
                perc_amarela = (dados['faixa_amarela'] / total * 100)
                perc_vermelha = (dados['faixa_vermelha'] / total * 100)
                perc_ocioso = (dados['motor_ocioso'] / total * 100)
                
                score = (perc_verde * 1.0) + (perc_azul * 0.8) - (perc_amarela * 0.5) - (perc_vermelha * 1.0) - (perc_ocioso * 0.3)
                
                evolucao_temporal.append({
                    'data': data.strftime('%Y-%m-%d'),
                    'score': round(score, 2),
                    'rpm_medio': round(sum(dados['rpms']) / len(dados['rpms']), 1) if dados['rpms'] else 0,
                    'velocidade_media': round(sum(dados['velocidades']) / len(dados['velocidades']), 1) if dados['velocidades'] else 0,
                    'perc_verde': round(perc_verde, 1),
                    'perc_vermelha': round(perc_vermelha, 1),
                    'perc_azul': round(perc_azul, 1),
                    'perc_amarela': round(perc_amarela, 1),
                    'perc_ocioso': round(perc_ocioso, 1),
                    'total_pontos': total
                })
    
    context = {
        'empresa_logada': empresa_logada,
        'motoristas_performance': motoristas_performance,
        'stats_gerais': stats_gerais,
        'evolucao_temporal': evolucao_temporal,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'periodo_dias': periodo_dias,
    }
    
    return render(request, 'umbrella360/performance_motoristas.html', context)

def jornada_motoristas(request):
    """View para análise de jornada de trabalho dos motoristas - tempo ao volante por dia"""
    from datetime import datetime, timedelta, date
    from collections import defaultdict
    import logging
    
    logger = logging.getLogger(__name__)
    

    # Verificar se há empresa logada
    empresa_logada_id = request.session.get('empresa_logada')
    empresa_logada = None
    
    if empresa_logada_id:
        try:
            empresa_logada = Empresa.objects.get(id=empresa_logada_id)
        except Empresa.DoesNotExist:
            return HttpResponseRedirect(reverse('login_view'))
    
    # Obter filtros de data
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    periodo_dias = request.GET.get('periodo_dias', '7')  # Padrão: últimos 7 dias
    
    # Buscar viagens eco relacionadas a motoristas
    from .models import Viagem_eco, Driver
    
    viagens_eco_query = Viagem_eco.objects.select_related('nome_motorista', 'unidade', 'unidade__empresa')
    viagens_eco_query = viagens_eco_query.filter(nome_motorista__isnull=False)
    
    if empresa_logada_id:
        viagens_eco_query = viagens_eco_query.filter(unidade__empresa_id=empresa_logada_id)
    
    # Aplicar filtros de data
    if data_inicio and data_fim:
        try:
            dt_inicio = datetime.strptime(data_inicio, '%Y-%m-%d')
            dt_fim = datetime.strptime(data_fim, '%Y-%m-%d')
            timestamp_inicio = int(dt_inicio.timestamp())
            timestamp_fim = int(dt_fim.replace(hour=23, minute=59, second=59).timestamp())
            viagens_eco_query = viagens_eco_query.filter(
                timestamp__gte=timestamp_inicio,
                timestamp__lte=timestamp_fim
            )
        except ValueError:
            pass
    elif periodo_dias and periodo_dias != 'todos':
        try:
            dias = int(periodo_dias)
            data_limite = datetime.now() - timedelta(days=dias)
            timestamp_limite = int(data_limite.timestamp())
            viagens_eco_query = viagens_eco_query.filter(timestamp__gte=timestamp_limite)
        except ValueError:
            pass
    
    # Ordenar por motorista e timestamp para análise sequencial
    viagens_eco_query = viagens_eco_query.order_by('nome_motorista_id', 'timestamp')
    
    # 🔥 OTIMIZAÇÃO: Limitar quantidade de dados processados
    # Se não há filtro de data, usar apenas últimos 7 dias por padrão
    if not data_inicio and not data_fim and (not periodo_dias or periodo_dias == 'todos'):
        periodo_dias = '7'
        dias = 7
        data_limite = datetime.now() - timedelta(days=dias)
        timestamp_limite = int(data_limite.timestamp())
        viagens_eco_query = viagens_eco_query.filter(timestamp__gte=timestamp_limite)
    
    # 🔥 LIMITE DE SEGURANÇA: Máximo 50.000 registros para evitar timeout
    total_registros = viagens_eco_query.count()
    if total_registros > 50000:
        # Pegar apenas os mais recentes
        viagens_eco_query = viagens_eco_query.order_by('-timestamp')[:50000]
    
    # Processar jornadas por motorista
    jornadas_motoristas = []
    
    if viagens_eco_query.exists():
        # 🔥 OTIMIZAÇÃO: Buscar apenas motoristas distintos primeiro
        motoristas_ids = list(viagens_eco_query.values_list('nome_motorista_id', flat=True).distinct()[:100])  # Limite de 100 motoristas
        
        # 🔥 OTIMIZAÇÃO: Carregar todos os motoristas de uma vez
        motoristas_dict = {m.id: m for m in Driver.objects.filter(id__in=motoristas_ids).select_related('empresa')}
        
        for motorista_id in motoristas_ids:
            try:
                motorista = motoristas_dict.get(motorista_id)
                if not motorista:
                    continue
                
                # 🔥 OTIMIZAÇÃO: Usar values() para reduzir overhead do ORM
                viagens_motorista = viagens_eco_query.filter(nome_motorista_id=motorista_id).values(
                    'timestamp', 'unidade_id', 'velocidade'
                ).order_by('timestamp')
                
                # Agrupar por dia e calcular tempo ao volante
                jornadas_por_dia = defaultdict(lambda: {
                    'timestamps': [],
                    'veiculos': set(),
                    'tempo_total_segundos': 0,
                    'velocidade_media': [],
                    'primeiro_ts': None,
                    'ultimo_ts': None
                })
                
                timestamps_anteriores = {}  # Por data
                
                for viagem in viagens_motorista:
                    timestamp_atual = viagem['timestamp']
                    data_viagem = date.fromtimestamp(timestamp_atual)
                    
                    # Adicionar veículo único
                    if viagem['unidade_id']:
                        jornadas_por_dia[data_viagem]['veiculos'].add(viagem['unidade_id'])
                    
                    # Coletar velocidade para média
                    if viagem['velocidade'] and viagem['velocidade'] > 0:
                        jornadas_por_dia[data_viagem]['velocidade_media'].append(viagem['velocidade'])
                    
                    # Registrar primeiro e último timestamp
                    if jornadas_por_dia[data_viagem]['primeiro_ts'] is None:
                        jornadas_por_dia[data_viagem]['primeiro_ts'] = timestamp_atual
                    jornadas_por_dia[data_viagem]['ultimo_ts'] = timestamp_atual
                    
                    # Calcular diferença de tempo com registro anterior DO MESMO DIA
                    if data_viagem in timestamps_anteriores:
                        timestamp_anterior = timestamps_anteriores[data_viagem]
                        diferenca_segundos = timestamp_atual - timestamp_anterior
                        
                        # Considerar como "sessão contínua" se diferença for <= 5 minutos (300 segundos)
                        if diferenca_segundos <= 300:
                            jornadas_por_dia[data_viagem]['tempo_total_segundos'] += diferenca_segundos
                    
                    timestamps_anteriores[data_viagem] = timestamp_atual
                    jornadas_por_dia[data_viagem]['timestamps'].append(timestamp_atual)
                
                # Calcular métricas agregadas por motorista
                total_dias_trabalhados = len(jornadas_por_dia)
                total_horas_trabalhadas = 0
                total_veiculos = set()
                jornadas_detalhadas = []
                
                for data_trab, dados in sorted(jornadas_por_dia.items()):
                    tempo_horas = dados['tempo_total_segundos'] / 3600
                    total_horas_trabalhadas += tempo_horas
                    total_veiculos.update(dados['veiculos'])
                    
                    # Determinar turno predominante usando timestamps já registrados
                    if dados['primeiro_ts']:
                        hora_inicio = datetime.fromtimestamp(dados['primeiro_ts']).hour
                        
                        if hora_inicio < 12:
                            turno = "Manhã"
                        elif hora_inicio < 18:
                            turno = "Tarde"
                        else:
                            turno = "Noite"
                        
                        jornada_inicio = datetime.fromtimestamp(dados['primeiro_ts'])
                        jornada_fim = datetime.fromtimestamp(dados['ultimo_ts'])
                    else:
                        turno = "N/A"
                        jornada_inicio = None
                        jornada_fim = None
                    
                    jornadas_detalhadas.append({
                        'data': data_trab,
                        'tempo_horas': round(tempo_horas, 2),
                        'tempo_formatado': f"{int(tempo_horas)}h {int((tempo_horas % 1) * 60)}min",
                        'veiculos_count': len(dados['veiculos']),
                        'sessoes_count': 0,  # Removido cálculo de sessões para otimização
                        'registros_count': len(dados['timestamps']),
                        'velocidade_media': round(sum(dados['velocidade_media']) / len(dados['velocidade_media']), 1) if dados['velocidade_media'] else 0,
                        'turno': turno,
                        'jornada_inicio': jornada_inicio,
                        'jornada_fim': jornada_fim,
                    })
                
                # Calcular médias
                media_horas_dia = total_horas_trabalhadas / total_dias_trabalhados if total_dias_trabalhados > 0 else 0
                
                # Classificar jornada
                if media_horas_dia >= 8:
                    classificacao = "Jornada Completa"
                    cor_classificacao = "#4CAF50"
                elif media_horas_dia >= 6:
                    classificacao = "Jornada Parcial"
                    cor_classificacao = "#FFC107"
                else:
                    classificacao = "Jornada Reduzida"
                    cor_classificacao = "#f44336"
                
                jornadas_motoristas.append({
                    'motorista': motorista,
                    'total_dias_trabalhados': total_dias_trabalhados,
                    'total_horas_trabalhadas': round(total_horas_trabalhadas, 2),
                    'media_horas_dia': round(media_horas_dia, 2),
                    'total_veiculos': len(total_veiculos),
                    'classificacao': classificacao,
                    'cor_classificacao': cor_classificacao,
                    'jornadas_detalhadas': sorted(jornadas_detalhadas, key=lambda x: x['data'], reverse=True),
                })
                
            except Exception as e:
                logger.error(f"Erro ao processar motorista {motorista_id}: {str(e)}")
                continue
    
    # Ordenar por total de horas trabalhadas (maior primeiro)
    jornadas_motoristas = sorted(jornadas_motoristas, key=lambda x: x['total_horas_trabalhadas'], reverse=True)
    
    # Estatísticas gerais
    if jornadas_motoristas:
        stats_gerais = {
            'total_motoristas': len(jornadas_motoristas),
            'total_horas_todas_jornadas': sum(m['total_horas_trabalhadas'] for m in jornadas_motoristas),
            'media_horas_por_motorista': sum(m['total_horas_trabalhadas'] for m in jornadas_motoristas) / len(jornadas_motoristas),
            'media_dias_trabalhados': sum(m['total_dias_trabalhados'] for m in jornadas_motoristas) / len(jornadas_motoristas),
            'motoristas_jornada_completa': sum(1 for m in jornadas_motoristas if m['classificacao'] == "Jornada Completa"),
            'motoristas_jornada_parcial': sum(1 for m in jornadas_motoristas if m['classificacao'] == "Jornada Parcial"),
            'motoristas_jornada_reduzida': sum(1 for m in jornadas_motoristas if m['classificacao'] == "Jornada Reduzida"),
        }
    else:
        stats_gerais = {
            'total_motoristas': 0,
            'total_horas_todas_jornadas': 0,
            'media_horas_por_motorista': 0,
            'media_dias_trabalhados': 0,
            'motoristas_jornada_completa': 0,
            'motoristas_jornada_parcial': 0,
            'motoristas_jornada_reduzida': 0,
        }
    
    # Preparar dados para gráfico de evolução temporal (horas por dia)
    evolucao_diaria = defaultdict(lambda: {'total_horas': 0, 'motoristas_ativos': set()})
    
    for motorista_data in jornadas_motoristas:
        for jornada_dia in motorista_data['jornadas_detalhadas']:
            data = jornada_dia['data']
            evolucao_diaria[data]['total_horas'] += jornada_dia['tempo_horas']
            evolucao_diaria[data]['motoristas_ativos'].add(motorista_data['motorista'].id)
    
    evolucao_temporal = [
        {
            #'data': data.strftime('%Y-%m-%d'),
            #'total_horas': round(dados['total_horas'], 1),
            #'motoristas_ativos': len(dados['motoristas_ativos']),
            #'media_horas': round(dados['total_horas'] / len(dados['motoristas_ativos']), 1) if dados['motoristas_ativos'] else 0
        }
    ]
    for data, dados in sorted(evolucao_diaria.items()):


        context = {
            'empresa_logada': empresa_logada,
            'jornadas_motoristas': jornadas_motoristas,
            'stats_gerais': stats_gerais,
            'evolucao_temporal': evolucao_temporal,
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'periodo_dias': periodo_dias,
        }
        
        logger.info(f"Jornada processada: {len(jornadas_motoristas)} motoristas, {len(evolucao_temporal)} dias")
        
        return render(request, 'umbrella360/Motoristas/jornada.html', context)
    
            
        # Retornar página com erro amigável
        context = {
            'empresa_logada': empresa_logada if 'empresa_logada' in locals() else None,
            'jornadas_motoristas': [],
            'stats_gerais': {
                'total_motoristas': 0,
                'total_horas_todas_jornadas': 0,
                'media_horas_por_motorista': 0,
                'media_dias_trabalhados': 0,
                'motoristas_jornada_completa': 0,
                'motoristas_jornada_parcial': 0,
                'motoristas_jornada_reduzida': 0,
            },
            'evolucao_temporal': [],
            'data_inicio': '',
            'data_fim': '',
            'periodo_dias': '7',
            'erro_mensagem': f'Erro ao processar dados: {str(e)}'
        }
        return render(request, 'umbrella360/Motoristas/jornada.html', context)
